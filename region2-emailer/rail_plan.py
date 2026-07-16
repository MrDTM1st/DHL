"""
Rail Plan builder (first cut).

Takes the raw CTMS "Short Rail Report" export (.csv) and produces the weekly
rail plan the way the SOP describes:
  * insert a "Leg" column (=1) after Trip Number
  * drop the columns that aren't used
  * drop manual / non-supplier orders (keep only the rail suppliers)
  * group the rows by DELIVERY DAY with a weekday separator row
  * write the master plan, then a plan per supplier-depot

    python rail_plan.py "<raw short_rail_rep .csv>"

Supplier-plan grouping is by collection depot (From Post Code), confirmed from
the real finished plans:
    DN16 1BP (Scunthorpe) -> British Steel
    DN6 0AA  (Askern)     -> Inframat / VAS   (Inf + VAS + the BS172943 code)
    SO40 4UT (Marchwood)  -> Arcelor Mittal
"""
import sys, os, csv, re, json
from datetime import datetime, date
from openpyxl.styles import PatternFill, Font, Alignment
import outbox

HERE = os.path.dirname(os.path.abspath(__file__))
try:
    RECIP = json.load(open(os.path.join(HERE, "_rail_recipients.json"), encoding="utf-8"))
except Exception:
    RECIP = {}

_COMMON = {"haulage", "transport", "transports", "services", "service", "rail", "freight",
           "logistics", "hire", "ltd", "limited", "the", "of", "and", "co", "uk", "group", "sons"}
_HAULIER_ALIAS = {"dhl northern operating centre": "dhl noc", "hhl": "hotspur hire (hhl)"}


def _core(s):
    return [t for t in re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).split()
            if len(t) >= 3 and t not in _COMMON]


def haulier_emails(carriers):
    """Match plan Carrier Names to the haulier contact list. Returns (emails,
    unmatched-carrier-names)."""
    hcfg = RECIP.get("hauliers", {})
    emails, unmatched = [], []
    for car in carriers:
        c = _HAULIER_ALIAS.get(str(car).strip().lower(), str(car).strip().lower())
        addrs = hcfg.get(c)
        if not addrs:
            for key, a in hcfg.items():
                if any(t in key.lower() for t in _core(c)):
                    addrs = a
                    break
        if addrs:
            emails += addrs
        elif "dhl" not in c:
            unmatched.append(car)
    return list(dict.fromkeys(emails)), unmatched

# Colours, matched from the real finished plans:
FILL_HEADER = PatternFill("solid", fgColor="FFFF00")    # header row - yellow
FILL_DAY = PatternFill("solid", fgColor="000000")       # day-separator bar - black
FONT_DAY = Font(bold=True, color="FFFFFF")              # ...with bold white text
FILL_NEW = PatternFill("solid", fgColor="92D050")       # new order (update) - green
FILL_CANCELLED = PatternFill("solid", fgColor="FF0000") # cancelled (update) - red
FILL_REG = PatternFill("solid", fgColor="FFC000")       # vehicle-reg update - orange
CENTER = Alignment(horizontal="center", vertical="center")


def is_adhoc(order_ref, product_desc, product_code):
    """Manual/adhoc rail orders are excluded from the rail plan. Caught by the
    'ADHOC' product, an AH-style adhoc order reference, or (main rule) a From
    Location that isn't one of the rail suppliers."""
    blob = f"{order_ref} {product_desc} {product_code}".lower()
    if "adhoc" in blob or "ad hoc" in blob:
        return True
    if re.match(r"\s*ah\d", str(order_ref or "").lower()):
        return True
    return False

DROP_COLS = {"del point ref", "product code", "collection date",
             "earliest coll time", "latest coll time", "instructions"}

# From Location prefix -> supplier (rows whose prefix isn't here are MANUAL/other and dropped)
SUPPLIER_PREFIX = [("bri", "British Steel"), ("bs", "British Steel"), ("vas", "VAS"),
                   ("ams", "Arcelor Mittal"), ("inf", "Inframat"), ("wm", "Whitemoor"),
                   ("whit", "Whitemoor")]

# collection depot (From Post Code prefix) -> (plan name, filename suffix)
def plan_for(from_loc, from_pc):
    pc = (from_pc or "").strip().upper().replace(" ", "")
    fl = (from_loc or "").strip().lower()
    if pc.startswith("DN161"):
        return ("British Steel", "BS")
    if pc.startswith("DN60"):
        return ("Inframat / VAS", "Inf VAS")
    if pc.startswith("SO40") or fl.startswith("ams"):
        return ("Arcelor Mittal", "AMS")
    return ("Other", "Other")


def supplier_of(from_loc):
    fl = (from_loc or "").strip().lower()
    for pre, name in SUPPLIER_PREFIX:
        if fl.startswith(pre):
            return name
    return None


def parse_del_date(v):
    v = str(v or "").strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(v, fmt).date()
        except Exception:
            pass
    return None


def parse_schedule(v):
    """Schedule is YYMMDD (260713 -> 13/07/2026) - the CTMS-intended delivery
    day, and the reliable key for the week + day grouping."""
    m = re.match(r"\s*(\d{2})(\d{2})(\d{2})\s*$", str(v or ""))
    if m:
        try:
            return date(2000 + int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            return None
    return None


def week_commencing(dates):
    ds = [d for d in dates if d]
    if not ds:
        return None
    monday = min(ds)
    monday = monday.fromordinal(monday.toordinal() - monday.weekday())  # back to Monday
    return monday


def load_raw(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    hdr = [h.strip() for h in rows[0]]
    return hdr, rows[1:]


def _refs_from_plan_xlsx(path):
    """Every Order Reference in a finished rail-plan .xlsx - skips the header,
    the black day-separator rows and blanks (they have no Order Reference)."""
    import openpyxl
    refs = set()
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return refs
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            hdr = next(rows)
        except StopIteration:
            continue
        col = next((i for i, h in enumerate(hdr or [])
                    if str(h or "").strip().lower() == "order reference"), None)
        if col is None:
            continue
        for row in rows:
            v = row[col] if col < len(row) else None
            if v is not None and str(v).strip():
                refs.add(str(v).strip())
    try:
        wb.close()
    except Exception:
        pass
    return refs


def previous_order_refs(wc):
    """Order References from the most recent rail-plan email in Outlook for this
    week-commencing - i.e. the latest version a colleague (or you) circulated.
    Returns a set, or None if no prior plan is found (so we don't green a whole
    fresh plan). Only used on a current-week UPDATE."""
    if wc is None:
        return None
    import tempfile, shutil
    wcs = wc.strftime("%d.%m")
    want = f"rail plan wc {wcs}"
    try:
        import build_drafts as bd
        ns = bd.get_ns()
        dhl = bd.dhl_store(ns)
        folders = [f for f in (bd.sub(dhl, "Inbox"), bd.sub(dhl, "Sent Items")) if f]
    except Exception:
        return None
    tmpdir = tempfile.mkdtemp(prefix="railprev_")
    best = None   # (when, [saved xlsx paths]) - the most recent matching email
    try:
        for fol in folders:
            try:
                items = fol.Items
                items.Sort("[ReceivedTime]", True)
            except Exception:
                continue
            n = 0
            for it in items:
                n += 1
                if n > 60:
                    break
                try:
                    subj = str(getattr(it, "Subject", "") or "").lower()
                    atts = it.Attachments
                    names = [str(atts.Item(j).FileName or "") for j in range(1, atts.Count + 1)]
                    if want not in subj and not any(want in nm.lower() for nm in names):
                        continue
                    rt = None
                    for attr in ("ReceivedTime", "SentOn", "LastModificationTime"):
                        v = getattr(it, attr, None)
                        if v is not None and 1990 < getattr(v, "year", 0) < 2100:
                            rt = v
                            break
                    when = datetime(rt.year, rt.month, rt.day, rt.hour, rt.minute) if rt else datetime.min
                    if best and when <= best[0]:
                        continue
                    paths = []
                    for j in range(1, atts.Count + 1):
                        att = atts.Item(j)
                        nm = str(att.FileName or "")
                        if nm.lower().endswith((".xlsx", ".xlsm")):
                            p = os.path.join(tmpdir, f"{n}_{nm}")
                            att.SaveAsFile(p)
                            paths.append(p)
                    if paths:
                        best = (when, paths)
                except Exception:
                    continue
        if not best:
            return None
        refs = set()
        for p in best[1]:
            refs |= _refs_from_plan_xlsx(p)
        return refs or None
    finally:
        try:
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass


def build(path, out_dir=None, update=False):
    hdr, data = load_raw(path)
    idx = {h.lower(): i for i, h in enumerate(hdr)}
    fl_i, fpc_i, dd_i, sch_i = idx["from location"], idx["from post code"], idx["delivery date"], idx["schedule"]
    car_i = idx.get("carrier name")

    # kept columns = original order, minus DROP_COLS, with Leg inserted after Trip Number
    keep_cols = [i for i, h in enumerate(hdr) if h.lower() not in DROP_COLS]
    out_hdr = []
    for i in keep_cols:
        out_hdr.append(hdr[i])
        if hdr[i].lower() == "trip number":
            out_hdr.append("Leg")

    def shape(row, sch=None):
        out = []
        for i in keep_cols:
            v = row[i] if i < len(row) else ""
            # correct a mis-planned delivery date to the day it's categorised under
            if sch is not None and i == dd_i and parse_del_date(v) != sch:
                v = sch.strftime("%d-%b-%y")
            out.append(v)
            if hdr[i].lower() == "trip number":
                out.append("1")
        return out

    ordref_i = idx["order reference"]
    pdesc_i, pcode_i = idx.get("product description"), idx.get("product code")
    kept, dropped, mism = [], [], []
    for r in data:
        if not any(c.strip() for c in r):
            continue
        oref = r[ordref_i] if ordref_i < len(r) else ""
        pdesc = r[pdesc_i] if pdesc_i is not None and pdesc_i < len(r) else ""
        pcode = r[pcode_i] if pcode_i is not None and pcode_i < len(r) else ""
        sup = supplier_of(r[fl_i])
        if is_adhoc(oref, pdesc, pcode):
            dropped.append((oref, "adhoc"))
            continue
        if sup is None:
            dropped.append((oref, f"not a rail supplier (From {r[fl_i]})"))
            continue
        sch = parse_schedule(r[sch_i])   # intended delivery day (reliable)
        dd = parse_del_date(r[dd_i])     # stated delivery date (check against sch)
        kept.append((r, sup, sch, dd))

    wc = week_commencing([sch for _, _, sch, _ in kept])
    # flag rows whose stated delivery date doesn't match the scheduled day (CTMS mis-plan)
    for r, sup, sch, dd in kept:
        if sch and (dd is None or dd != sch):
            mism.append((r[idx["order reference"]], r[sch_i], r[dd_i]))

    # UPDATE (current-week) mode: pull the previous plan a colleague circulated
    # for this week from Outlook, so we can highlight the NEW manifests green.
    prev_refs = previous_order_refs(wc) if update else None

    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    def write_plan(rows, suffix):
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = f"wc {wc.strftime('%d.%m')}" if wc else "Rail Plan"
        ws.append(out_hdr)
        for cell in ws[1]:                       # header row - yellow
            cell.fill = FILL_HEADER
        # group by delivery-day weekday, in Mon..Sun order; unknown/mismatch last
        by_day, carriers = {}, set()
        for r, sup, sch, dd in rows:
            k = sch.weekday() if sch else 99
            by_day.setdefault(k, []).append((r, sch))
            if car_i is not None and car_i < len(r) and str(r[car_i] or "").strip():
                carriers.add(str(r[car_i]).strip())
        for k in sorted(by_day):
            ws.append([DAYS[k] if k < 7 else "CHECK - date mismatch"] + [""] * (len(out_hdr) - 1))
            for cell in ws[ws.max_row]:          # day-separator bar - black w/ bold white text
                cell.fill = FILL_DAY
                cell.font = FONT_DAY
            for r, sch in by_day[k]:
                ws.append(shape(r, sch))
                # new manifest since the previous plan -> highlight the row green
                if update and prev_refs:
                    ref = str(r[ordref_i]).strip() if ordref_i < len(r) else ""
                    if ref and ref not in prev_refs:
                        for cell in ws[ws.max_row]:
                            cell.fill = FILL_NEW
        # centre every cell, and size each column to its content
        for row in ws.iter_rows():
            for c in row:
                c.alignment = CENTER
        for col in ws.columns:
            mx = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, 10), 55)
        wcs = wc.strftime("%d.%m") if wc else "TBC"
        name = f"Rail Plan wc {wcs} v1" + (f" {suffix}" if suffix else "") + ".xlsx"
        p = os.path.join(out_dir, name) if out_dir else outbox.path(name)
        wb.save(p)
        return p, name, sum(len(v) for v in by_day.values()), carriers

    # master = all kept
    mpath, mfile, mcount, mcarr = write_plan(kept, "")
    # per supplier-plan
    plans = {}
    for r, sup, sch, dd in kept:
        pn, sfx = plan_for(r[fl_i], r[fpc_i])
        plans.setdefault((pn, sfx), []).append((r, sup, sch, dd))
    made = [{"path": mpath, "name": mfile, "count": mcount, "chain": None, "suffix": "", "carriers": mcarr}]
    for (pn, sfx), rows in plans.items():
        p, f, c, carr = write_plan(rows, sfx)
        made.append({"path": p, "name": f, "count": c, "chain": pn, "suffix": sfx, "carriers": carr})

    print(f"Week commencing: {wc}  |  kept {len(kept)} rows, dropped {len(dropped)} manual/non-supplier")
    if update:
        wcs = wc.strftime("%d.%m") if wc else "TBC"
        if prev_refs is None:
            print(f"  UPDATE: no previous rail plan found in Outlook for wc {wcs} - nothing highlighted green.")
        else:
            cur = {str(r[ordref_i]).strip() for r, _, _, _ in kept if ordref_i < len(r) and str(r[ordref_i]).strip()}
            newr = sorted(cur - prev_refs)
            print(f"  UPDATE: {len(newr)} new manifest(s) highlighted green"
                  + ((": " + ", ".join(newr)) if newr else " (none new)."))
    if dropped:
        print("  EXCLUDED (adhoc / manual / not a rail supplier):")
        for o, reason in dropped:
            print(f"     {o}  ({reason})")
    if mism:
        print(f"  ~ {len(mism)} row(s) had a wrong delivery date - corrected to the scheduled day:")
        for o, sch, dd in mism:
            print(f"     {o}  was {dd}  ->  {sch}")
    print("  FILES:")
    for m in made:
        print(f"     {m['count']:3} rows  {(m['chain'] or 'MASTER'):16}  {m['name']}")
    return wc, made, mism


def recipients_for(m):
    """To-list for one plan file: its supplier chain (per plan_to_chain) +
    always the DHL colleagues + the hauliers actually on that plan."""
    chain_key = RECIP.get("_plan_to_chain", {}).get(m["suffix"])
    to = list(RECIP.get("chains", {}).get(chain_key, [])) if chain_key else []
    to += RECIP.get("dhl_always", [])
    hem, unmatched = haulier_emails(m["carriers"])
    to += hem
    return list(dict.fromkeys(x for x in to if x)), unmatched


def sendoff(path, send=False, update=False):
    """Build the plans then draft/send one email per plan to its recipients.
    Plans are written to the outbox so they're downloadable from the dashboard.
    update=True (current-week) greens the new manifests vs the previous plan."""
    wc, made, mism = build(path, update=update)
    wcs = wc.strftime("%d.%m") if wc else "TBC"
    outlook = acct = ns = None
    if send:
        import win32com.client, build_drafts as bd, send_order as so
        ns = bd.get_ns(); acct = so.dhl_account(ns)
        outlook = win32com.client.Dispatch("Outlook.Application")
    print(f"\n=== RAIL PLAN SEND-OFF wc {wcs}  ({'SENDING' if send else 'PREVIEW - nothing sent'}) ===")
    for m in made:
        to, unmatched = recipients_for(m)
        label = m["chain"] or "MASTER"
        subj = f"Rail Plan wc {wcs}" + (f" - {m['chain']}" if m["chain"] else "")
        print(f"\n[{label}]  attach: {m['name']}  ->  {len(to)} recipients")
        print("   TO: " + "; ".join(to))
        if unmatched:
            print("   !! carrier(s) not matched to a haulier email: " + ", ".join(unmatched))
        if send:
            import build_drafts as bd, send_order as so
            mail = outlook.CreateItem(0)
            to_str, _cc, _rm = bd.clean_to_cc("; ".join(to))
            mail.To = to_str
            mail.Subject = subj
            message = (f"Hi all,\n\nPlease find attached the rail plan for week commencing "
                       f"{wc.strftime('%d/%m/%Y')}"
                       + (f" - {m['chain']}." if m["chain"] else ".")
                       + "\n\nMany thanks.")
            bd._attach_qr(mail)                          # inline QR for the signature
            mail.HTMLBody = bd.html_from_message(message)  # message + your signature
            mail.Attachments.Add(m["path"])
            if not so.bind_account(mail, acct):
                print("   ABORT: could not bind DHL account - NOT sent"); continue
            mail.Send()
            print("   SENT")
    if send and ns is not None:
        try:
            ns.SendAndReceive(False)
        except Exception:
            pass


if __name__ == "__main__":
    a = sys.argv[1:]
    update = "--update" in a                      # current-week update: green the new manifests
    a = [x for x in a if x != "--update"]
    if a and a[0] == "send":
        if len(a) < 2 or not os.path.exists(a[1]):
            print("Usage: python rail_plan.py send <raw csv> [go] [--update]"); sys.exit(1)
        sendoff(a[1], send=(len(a) > 2 and a[2].lower() == "go"), update=update)
    elif a and os.path.exists(a[0]):
        build(a[0], update=update)
    else:
        print("Usage: python rail_plan.py <raw csv> [--update]   |   "
              "python rail_plan.py send <raw csv> [go] [--update]")
