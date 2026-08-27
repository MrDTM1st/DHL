"""
Seasonal treatments: a Seasonal Treatment Order -> the NR upload CSV, plus a
cover request to the haulier that site is allocated to.

These arrive from DHL Road (forwarded by Darren Knightingale) as one
"Seasonal Treatment Order v<version>.xlsx" per job. The season is the autumn
railhead-treatment programme and the winter one - Track Grip 60, Electra Gel,
Kilfrost, Arrow and their empties/part-fulls.

The order form already carries a filled "RHPC Upload" sheet, and that sheet is
the 47-column order dict nr_csv.transform consumes, column for column. So this
does NOT re-derive anything from the human-facing Order Form: it reads the
sheet the form itself produced, which is the one the requester checked. The
only fix-up is the four time columns, which the form spells with underscores
(collection_time) and nr_csv spells with spaces (collection time).

Who gets asked to cover is NOT in the order. It is in the separate allocation
sheet - "Seasonal sites with haulier allocation <years>.xlsx" - which pairs
every DELIVERY site with a Haulier 1 and a Haulier 2. Import it once per
season and it is cached next to this file:

    python seasonal.py import "<Seasonal sites with haulier allocation.xlsx>"
    python seasonal.py "<Seasonal Treatment Order.xlsx>" [more orders...]

Only Haulier 1 is asked. Haulier 2 is recorded so it is there to escalate to
by hand, but the tool never emails it - two hauliers quoting the same job at
once is how you end up paying for both.

Sends nothing. Writes the CSV into the outbox and stages the cover requests
for Review & send, the same as every other route here.
"""
import os
import sys
import json
import datetime

import nr_csv
import outbox

HERE = os.path.dirname(os.path.abspath(__file__))
SITES = os.path.join(HERE, "_seasonal_sites.json")
PENDING = os.path.join(HERE, "_pending_email.json")

UPLOAD_SHEET = "RHPC Upload"

# The order form writes these with underscores; nr_csv reads them with spaces.
# Same column, two spellings - map rather than teach nr_csv a second name.
TIME_COLS = {
    "collection_time": "collection time",
    "collection_time_end": "collection time end",
    "delivery_time": "delivery time",
    "delivery_time_end": "delivery time end",
}


# ---------- the allocation sheet ----------
def _norm_pc(pc):
    """Postcodes are the join key, so they must compare on shape not spacing."""
    return "".join(str(pc or "").upper().split())


def import_allocation(path):
    """Cache the season's site -> haulier allocation next to this file.

    Keyed on the delivery POSTCODE. Site names differ between the allocation
    sheet and the order form for the same place - the sheet says "DBC Wigan",
    the order says "Wigan" - so matching on the name alone would miss. The
    postcode is written the same in both.
    """
    import openpyxl
    import warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sites"] if "Sites" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # Row 1 is the sheet's title; row 2 carries the headers, after a blank
    # column A. Find it rather than assume, so a re-issued sheet that gains a
    # row at the top does not silently shift every field by one.
    head = None
    for i, r in enumerate(rows[:8]):
        vals = [str(c).strip().lower() for c in r if c is not None]
        if "site name" in vals and "post code" in vals:
            head = i
            break
    if head is None:
        raise SystemExit(f"{os.path.basename(path)}: no header row with "
                         f"'Site Name' and 'Post Code' - is this the allocation sheet?")
    hdr = [str(c).strip().lower() if c is not None else "" for c in rows[head]]

    def col(name):
        return hdr.index(name) if name in hdr else None

    ix = {k: col(k) for k in ("product", "supplier", "site name", "post code",
                              "loading", "vehicle type", "haulier 1", "haulier 2")}
    missing = [k for k, v in ix.items() if v is None]
    if missing:
        raise SystemExit(f"{os.path.basename(path)}: allocation sheet is missing "
                         f"column(s): {', '.join(missing)}")

    def g(r, key):
        i = ix[key]
        v = r[i] if i is not None and i < len(r) else None
        return "" if v is None else str(v).strip()

    sites = {}
    for r in rows[head + 1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        pc = _norm_pc(g(r, "post code"))
        name = g(r, "site name")
        if not pc or not name:
            continue
        sites[pc] = {
            "site": name, "postcode": g(r, "post code"),
            "product": g(r, "product"), "supplier": g(r, "supplier"),
            "loading": g(r, "loading"), "vehicle": g(r, "vehicle type"),
            "haulier1": g(r, "haulier 1"), "haulier2": g(r, "haulier 2"),
        }
    with open(SITES, "w", encoding="utf-8") as f:
        json.dump({"source": os.path.basename(path), "sites": sites}, f, indent=1)
    allocated = sum(1 for s in sites.values() if s["haulier1"])
    return sites, allocated


def load_allocation():
    try:
        return json.load(open(SITES, encoding="utf-8")).get("sites", {})
    except Exception:
        return {}


# ---------- haulier lookup ----------
def resolve_haulier(code):
    """An allocation code ("NOC", "Revis") -> the directory record.

    The allocation sheet writes hauliers as the yard shorthand, the contact
    list writes them out in full ("Revis" / "Revis of York"), so this matches
    on the shorthand being contained in the full name. Returns None rather
    than guessing when nothing matches - an unrecognised code must surface,
    not quietly pick the nearest-looking haulier.
    """
    code = str(code or "").strip()
    if not code:
        return None
    try:
        import hauliers
        d = hauliers.load()
    except Exception:
        return None
    every = []
    for grp in ("hauliers", "couriers", "services"):
        every.extend(d.get(grp, []))

    def squash(s):
        # Letters and digits only. Spaces matter here: "DEO" is D E O'Reilly,
        # whose name has the initials spaced out, so anything that keeps
        # spaces or apostrophes fails to match it and the site silently gets
        # no haulier.
        return "".join(c for c in str(s).lower() if c.isalnum())

    key = squash(code).rstrip("s")
    best = None
    for h in every:
        name = squash(h.get("name", ""))
        if key and key in name:
            # prefer the shortest matching name - "Revis" should land on
            # "Revis of York", not on something that merely contains it
            if best is None or len(name) < len(squash(best.get("name", ""))):
                best = h
    return best


# ---------- the order ----------
def _fmt(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%d/%m/%Y %H:%M")
    if isinstance(v, datetime.date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    return "" if v is None else str(v).strip()


def read_order(path):
    """The RHPC Upload sheet's rows, as order dicts nr_csv can transform."""
    import openpyxl
    import warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)
    if UPLOAD_SHEET not in wb.sheetnames:
        wb.close()
        raise SystemExit(
            f"{os.path.basename(path)}: no '{UPLOAD_SHEET}' sheet - this does not "
            f"look like a Seasonal Treatment Order.\n  sheets found: "
            f"{', '.join(wb.sheetnames)}")
    ws = wb[UPLOAD_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        o = {}
        for i, h in enumerate(hdr):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            o[TIME_COLS.get(h, h)] = _fmt(v) if isinstance(
                v, (datetime.datetime, datetime.date, datetime.time)) else v
        # a row with no order ref is a spare template row, not a job
        if str(o.get("Customer Order No") or "").strip():
            out.append(o)
    return out


# ---------- the cover request ----------
def window(start, end):
    """"05/09/2026 08:00 16:00" - the way the send-outs already read.

    The date is written once. It reappears on the end only when the window
    actually crosses midnight, which is the one case where dropping it would
    turn a 22:30-00:30 night drop into something that looks like it finishes
    22 hours early.
    """
    s, e = str(start or "").strip(), str(end or "").strip()
    if not s:
        return e
    if not e:
        return s
    sd, _, st = s.partition(" ")
    ed, _, et = e.partition(" ")
    if sd == ed and st and et:
        return f"{sd} {st} {et}"
    return f"{s} {e}"


def _offload(o):
    v = str(o.get("HIAB") or "").strip()
    return v if v and v.upper() != "N" else "none required"


def cover_request(orders, code, haulier, source):
    """One "would you be able to cover" email for one haulier.

    Laid out exactly like the send-outs already in Sent Items - Order /
    Collection / Delivery / dates / Materials / Vehicle / Offloading - so a
    haulier who has had one before reads it without re-learning it.
    """
    lines = ["Hi,", "",
             f"Would you be able to cover the job{'s' if len(orders) != 1 else ''} below;"]
    for o in orders:
        g = lambda k: str(o.get(k) or "").strip()          # noqa: E731
        qty, prod = g("Product Qty"), g("Product / Service Code")
        lines += [
            "",
            f"Order: {g('Customer Order No')}",
            f"Collection: {g('Site Name - Collection')}, {g('Postcode')}",
            f"Collection date/time: {window(g('collection time'), g('collection time end'))}",
            f"Delivery: {g('Delivery Point')}, {g('D Postcode')}",
            f"Delivery date/time: {window(g('delivery time'), g('delivery time end'))}",
            f"Materials: {qty}x {prod}" if qty or prod else "Materials:",
            f"Vehicle: {g('Vehicle Type')}",
            f"Offloading: {_offload(o)}",
        ]
    lines += ["", "Let me know if you can cover and what the cost would be."]

    emails = haulier.get("emails") if haulier else []
    dates = [str(o.get("delivery time") or "")[:10] for o in orders if o.get("delivery time")]
    return {
        "to": "; ".join(emails or []), "cc": "", "name": "",
        "subject": (f"Seasonal treatment {orders[0].get('Customer Order No')}"
                    + (f" +{len(orders) - 1} more" if len(orders) > 1 else "")
                    + " - can you cover?"),
        "message": "\n".join(lines),
        "date": dates[0] if dates else "",
        "area": "", "orders": [str(o.get("Customer Order No")) for o in orders],
        "product_codes": sorted({str(o.get("Product / Service Code") or "")
                                 for o in orders if o.get("Product / Service Code")}),
        "materials": ", ".join(f"{o.get('Product Qty')}x {o.get('Product / Service Code')}"
                               for o in orders),
        "site": str(orders[0].get("Delivery Point") or ""),
        "postcode": str(orders[0].get("D Postcode") or ""),
        "source": source, "attach": [],
        # a haulier cover request, NOT a delivery contact - see send_emails
        "no_track": True, "haulier": code.upper(), "_metric": "haulier_request",
    }


def is_internal(haulier):
    """DHL's own people, who are spoken to on Teams rather than emailed.

    Decided on the addresses, not the name: anyone whose every contact is
    @dhl.com is a colleague, and "would you be able to cover, and what would
    it cost" is the wrong thing to send a colleague. NOC is the one this
    exists for, but the rule holds for any internal allocation.
    """
    emails = [str(e).strip().lower() for e in (haulier or {}).get("emails") or []]
    return bool(emails) and all(e.endswith("@dhl.com") for e in emails)


def teams_message(orders, haulier):
    """The same ask, laid out to be pasted straight into a Teams chat.

    No signature and no greeting block - a Teams message carries the sender
    already, and the email footer pasted into a chat window is noise.
    """
    name = str((haulier or {}).get("name") or "").replace("DHL ", "")
    lines = [f"Hi{' ' + name if name else ''}, can you cover "
             f"{'these' if len(orders) != 1 else 'this'} seasonal treatment "
             f"run{'s' if len(orders) != 1 else ''}?", ""]
    for o in orders:
        g = lambda k: str(o.get(k) or "").strip()          # noqa: E731
        off = _offload(o)
        lines += [
            g("Customer Order No"),
            f"Collect: {g('Site Name - Collection')}, {g('Postcode')} - "
            f"{window(g('collection time'), g('collection time end'))}",
            f"Deliver: {g('Delivery Point')}, {g('D Postcode')} - "
            f"{window(g('delivery time'), g('delivery time end'))}",
            f"{g('Product Qty')}x {g('Product / Service Code')}, {g('Vehicle Type')}"
            + (", no offload" if off == "none required" else f", offload {off}"),
            "",
        ]
    lines.append("Let me know if you can take "
                 f"{'them' if len(orders) != 1 else 'it'} and I will get "
                 f"{'them' if len(orders) != 1 else 'it'} raised.")
    return "\n".join(lines)


def main():
    args = [a for a in sys.argv[1:] if a.strip()]
    if args and args[0].lower() == "import":
        if len(args) < 2 or not os.path.exists(args[1]):
            print("Give me the allocation sheet: "
                  "python seasonal.py import \"<...allocation....xlsx>\"")
            return 2
        sites, allocated = import_allocation(args[1])
        print(f"Imported {len(sites)} seasonal site(s) from "
              f"{os.path.basename(args[1])} - {allocated} with a Haulier 1.")
        print(f"  cached: {SITES}")
        return 0

    paths = [a for a in args if os.path.exists(a)]
    if not paths:
        print(__doc__)
        print("SEASONAL_RESULT orders=0")
        return 2

    alloc = load_allocation()
    if not alloc:
        print("!! No allocation cached yet - I do not know which haulier covers "
              "which site.\n   Run:  python seasonal.py import "
              "\"<Seasonal sites with haulier allocation.xlsx>\"")
        print("SEASONAL_RESULT orders=0")
        return 2

    orders, warnings_ = [], []
    for p in paths:
        got = read_order(p)
        if not got:
            warnings_.append(f"{os.path.basename(p)}: no rows on the "
                             f"'{UPLOAD_SHEET}' sheet")
        for o in got:
            o["_source"] = os.path.basename(p)
        orders.extend(got)
    if not orders:
        for w in warnings_:
            print(f"  !! {w}")
        print("SEASONAL_RESULT orders=0")
        return 1

    stamp = datetime.datetime.now().strftime("%d%m%Y%H%M%S")
    print(f"Seasonal treatments - {len(orders)} order(s) from "
          f"{len(paths)} file(s)\n")

    # ---- match each delivery site to its allocated haulier ----
    by_haulier, unmatched = {}, []
    for o in orders:
        pc = _norm_pc(o.get("D Postcode"))
        site = alloc.get(pc)
        code = (site or {}).get("haulier1", "")
        rec = resolve_haulier(code) if code else None
        note = ""
        if site is None:
            note = "site not on the allocation sheet"
            unmatched.append(o)
        elif not code:
            note = f"no Haulier 1 allocated for {site['site']}"
            unmatched.append(o)
        elif rec is None:
            note = f"haulier '{code}' is not in the contact list"
            unmatched.append(o)
        elif not rec.get("emails"):
            note = f"{rec.get('name')} has no email in the contact list"
            unmatched.append(o)
        else:
            by_haulier.setdefault(code, {"rec": rec, "orders": []})
            by_haulier[code]["orders"].append(o)
            if rec.get("do_not_use"):
                note = f"!! {rec.get('name')} is marked DO NOT USE in the contact list"
        print(f"  {str(o.get('Customer Order No')):16} "
              f"{str(o.get('Site Name - Collection'))[:20]:20} {str(o.get('Postcode')):9}"
              f" -> {str(o.get('Delivery Point'))[:14]:14} {str(o.get('D Postcode')):9}"
              f" {(code or '-'):9} {note}")
        if site and site.get("haulier2"):
            o["_haulier2"] = site["haulier2"]

    # ---- the CSV ----
    records = nr_csv.transform(orders)
    csv_out = outbox.path(f"NR_seasonal_{stamp}.csv")
    nr_csv.write_csv(records, csv_out)
    print(f"\n  CSV: {csv_out}")

    # ---- stage one cover request per haulier ----
    # Internal allocations (NOC) are spoken to on Teams, so they get a block of
    # text to paste rather than a staged email. It goes in the outbox as a .txt
    # so it lands on the Files card with everything else, and is printed here
    # so a command-line run can be copied straight out of the console.
    staged, teams = [], []
    for code, d in sorted(by_haulier.items()):
        src = ", ".join(sorted({o.get("_source", "") for o in d["orders"]}))
        if is_internal(d["rec"]):
            teams.append((code, d["rec"], teams_message(d["orders"], d["rec"])))
        else:
            staged.append(cover_request(d["orders"], code, d["rec"], src))
    # Printed between markers, NOT written to a file. The message is something
    # you copy off the screen once and paste into Teams - saving it to the
    # outbox would put a throwaway .txt on the Files card next to the uploads
    # that actually matter, and push a real one off the twelve it keeps.
    # The markers are what the agent parses to fill the dashboard panel; the
    # block between them is the message exactly, no indent to strip.
    for code, rec, msg in teams:
        print(f"\n  TEAMS: {code} ({rec.get('name')}) is internal - no email "
              f"staged. Paste this into Teams:\n")
        print(f"TEAMS_BEGIN {code} ({rec.get('name')})")
        print(msg)
        print("TEAMS_END")
    if staged:
        # MERGE, do not replace. Seasonal orders arrive one file at a time, so
        # a straight overwrite would mean processing the second order silently
        # destroys the first one's cover request - the email is gone and
        # nothing says so. Anything already staged for the same haulier and the
        # same orders is replaced (a re-run of the same file should not stage
        # it twice); everything else is left alone.
        try:
            existing = json.load(open(PENDING, encoding="utf-8"))
            if not isinstance(existing, list):
                existing = []
        except Exception:
            existing = []
        mine = {(e["haulier"], tuple(sorted(e["orders"]))) for e in staged}
        kept = [e for e in existing
                if (e.get("haulier"), tuple(sorted(e.get("orders") or []))) not in mine]
        with open(PENDING, "w", encoding="utf-8") as f:
            json.dump(kept + staged, f, indent=1)
        if kept:
            print(f"  (kept {len(kept)} email(s) already staged)")
        for e in staged:
            print(f"  COVER: {e['haulier']:9} -> {e['to']}  "
                  f"({len(e['orders'])} order(s)) ready for Review & send")
    for w in warnings_:
        print(f"  !! {w}")
    if unmatched:
        print(f"  !! {len(unmatched)} order(s) have no haulier to ask - "
              f"they are ON the CSV but nobody has been emailed:")
        for o in unmatched:
            print(f"       {o.get('Customer Order No')}  "
                  f"{o.get('Delivery Point')} {o.get('D Postcode')}")
    print(f"\n  COVER_READY {1 if staged else 0}")
    print(f"SEASONAL_RESULT orders={len(orders)} hauliers={len(by_haulier)} "
          f"unmatched={len(unmatched)}")
    try:
        import metrics
        metrics.log("seasonal_built",
                    orders=[str(o.get("Customer Order No")) for o in orders],
                    what=f"{len(by_haulier)} haulier(s)")
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    sys.exit(main())
