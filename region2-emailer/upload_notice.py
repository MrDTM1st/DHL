"""
The email that goes out AFTER a Synergy upload - previewed, never sent.

Uploading the extract is only half the job. The other half is telling people:
the materials team that owns the product ("Ballast done" to Track Aggregates),
the rest of the planners so nobody uploads the same extract twice, and - the
part that actually gets missed - the planner who owns another region when one
of their orders is running out of time.

That last one is why this exists. The upload CSV is built from every region's
orders, but only Region 2 is watched by this toolkit. An R4 collection due
tomorrow sits in the sheet looking exactly like an R4 collection due in three
weeks, and the person who has to book a vehicle for it never finds out until
it is late. So each order is put through nr_regions, the week's Transport
Planner sheet says who is covering that region on that day, and anything
inside the urgency window is called out by name.

    _planner_sheet.xlsx   the cached copy of the week's rota
    _pending_email.json   the email, ready for Review & send

Nothing here sends. The email is staged in exactly the shape date_query and
send_order.py sendjson already use, so it lands in the same Review & send panel
as everything else and a human presses the button.
"""
import json
import os
import re
from datetime import datetime, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
SHEET_CACHE = os.path.join(HERE, "_planner_sheet.xlsx")
PENDING = os.path.join(HERE, "_pending_email.json")

# How close a collection has to be before another region's planner is told.
# Three days is what gets called out by hand today; anything further away is
# ordinary work and naming it would train people to ignore the whole section.
URGENT_DAYS = 3

# Which row of the rota covers the job. Ballast rides in a tipper and
# everything else rides on a flat, so the product decides which of the two
# rows for that region is the one with the right name in it.
TIPPER_PRODUCTS = ("ballast", "loose ballast")


# ---------- the rota ----------
def fetch_sheet(save_to=SHEET_CACHE):
    """Pull the newest "Transport Planners wc ....xlsx" out of Outlook.

    The rota arrives as an attachment every week and is never on disk, so the
    only copy is in the mailbox. A cached copy is kept next to this file: a
    stale rota still names the right people far more often than no rota at all,
    and the alternative is the urgency section silently disappearing on any day
    Outlook is not reachable.
    """
    try:
        import win32com.client
    except Exception:
        return save_to if os.path.exists(save_to) else ""
    try:
        app = win32com.client.Dispatch("Outlook.Application")
        ns = app.GetNamespace("MAPI")
        # NOT GetDefaultFolder on the namespace - the profile's default store is
        # a personal account, and reading it finds no rota at all. Same trap as
        # check_sending._sent_folder.
        acct = None
        import build_drafts as bd
        want = bd.DHL_SMTP.lower()
        for i in range(1, ns.Accounts.Count + 1):
            a = ns.Accounts.Item(i)
            if str(getattr(a, "SmtpAddress", "")).lower() == want:
                acct = a
                break
        if acct is None:
            return save_to if os.path.exists(save_to) else ""
        items = acct.DeliveryStore.GetDefaultFolder(6).Items
        items.Sort("[ReceivedTime]", True)
        for n, m in enumerate(items):
            if n > 400:
                break
            try:
                if "transport planner" not in str(m.Subject or "").lower():
                    continue
                for k in range(1, m.Attachments.Count + 1):
                    att = m.Attachments.Item(k)
                    fn = str(att.FileName or "")
                    if fn.lower().endswith((".xlsx", ".xlsm")) and "planner" in fn.lower():
                        att.SaveAsFile(save_to)
                        return save_to
            except Exception:
                continue
    except Exception:
        pass
    return save_to if os.path.exists(save_to) else ""


def _cell_date(v):
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    if hasattr(v, "year"):
        return datetime(v.year, v.month, v.day)
    return None


def parse_rota(path):
    """Just the region cover - see parse_sheet, which this wraps."""
    return parse_sheet(path)[0]


def parse_sheet(path):
    """(rota, people) from the Main sheet.

    rota is {"R2": {date: {"flats": [names], "tippers": [names]}}}, keyed on
    the real date from row 1 rather than a weekday name: the sheet is one week
    per file, and matching a job's date against the sheet's own dates is the
    only way to know the rota being read actually covers that day.

    people is every planner named ANYWHERE on the sheet. The region rows alone
    are too narrow for the To line - the same team also covers DHL Road shifts,
    Projects and Ad Hoc, and they are on the real sends. Anyone whose name is
    on the week's sheet is working that week and wants to know an upload ran.
    """
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Main"] if "Main" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # Row 1 carries the dates across the columns. Find them rather than assume
    # a fixed offset - a re-issued sheet that gains a column shifts every name.
    col_date = {}
    for r in rows[:4]:
        for i, c in enumerate(r or []):
            d = _cell_date(c)
            if d:
                col_date[i] = d
        if col_date:
            break

    known = set(team_emails()[0])
    out, people = {}, []
    for r in rows:
        # Anyone on the sheet who is on the roster is working this week.
        for c in (r or []):
            for n in _names(c):
                k = n.lower()
                if k in known and k not in people:
                    people.append(k)
        label = str((r or [None])[0] or "").strip()
        m = re.match(r"(?i)region\s*([1-4])", label)
        if not m:
            continue
        region = "R" + m.group(1)
        kind = "tippers" if re.search(r"(?i)tipper", label) else "flats"
        for i, d in col_date.items():
            names = _names(r[i] if i < len(r) else None)
            if names:
                out.setdefault(region, {}).setdefault(d, {}).setdefault(kind, [])
                out[region][d][kind] = names
    return out, people


def _names(cell):
    """"Zoe / Delali" is two people covering one region that day, not a name."""
    txt = str(cell or "").strip()
    if not txt or txt.upper() in ("H", "BH"):
        return []
    return [p.strip() for p in re.split(r"[/,&]| and ", txt) if p.strip()]


# ---------- people ----------
def team_emails():
    """First name -> work address, from the team roster the rest of the
    toolkit already uses. The rota writes first names only."""
    try:
        import build_drafts as bd
        cfg = bd.team_config()
    except Exception:
        try:
            cfg = json.load(open(os.path.join(HERE, "config", "team.json"), encoding="utf-8"))
        except Exception:
            cfg = {}
    by_first, me = {}, str(cfg.get("me") or "").lower()
    for mem in cfg.get("members", []):
        name = str(mem.get("name") or "").strip()
        email = str(mem.get("email") or "").strip()
        if name and email:
            by_first[name.split()[0].lower()] = email
    return by_first, me


# ---------- products ----------
def material_of(desc):
    """The product family, using the one classifier the toolkit already has."""
    try:
        import build_drafts as bd
        return bd.product_type(desc)
    except Exception:
        d = str(desc or "").upper()
        for key, label in (("SLEEPER", "sleepers"), ("BALLAST", "ballast"),
                           ("RAIL", "rails")):
            if key in d:
                return label
        return ""


def _summary(kinds):
    """"Ballast done" is what these emails actually say. Loose and bagged
    ballast are two products to a planner but one word to a materials team, so
    they collapse here - "Loose Ballast, Ballast done" reads like a mistake."""
    order, seen = [], set()
    for k in kinds:
        k = "ballast" if "ballast" in str(k).lower() else str(k).lower()
        if k and k not in seen:
            seen.add(k)
            order.append(k)
    if not order:
        return "Orders"
    if len(order) == 1:
        return order[0].title()
    return (", ".join(x.title() for x in order[:-1]) + " and " + order[-1].title())


def material_team(kind):
    """The Network Rail team that owns this product, or None when the product
    does not map to one - S&C and pads have no team on the list, and guessing a
    recipient is worse than leaving the Cc for a human to fill in."""
    try:
        import hauliers
        teams = hauliers.MATERIALS_TEAMS
    except Exception:
        return None
    k = str(kind or "").lower()
    if "ballast" in k:
        return teams.get("ballast")
    if "sleeper" in k:
        return teams.get("sleepers")
    if "rail" in k:
        return teams.get("rails")
    return None


# ---------- urgency ----------
def _when(order):
    """The collection is the deadline that bites - it is the end somebody has
    to put a vehicle against - so urgency is measured from it, falling back to
    the delivery when the extract carries no collection time."""
    for key in ("collection time", "delivery time"):
        raw = str(order.get(key) or "").strip()
        if not raw:
            continue
        for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw[:16] if " " in raw else raw, fmt)
            except Exception:
                continue
    return None


def urgency(order, today=None):
    """(days, phrase) for anything inside the window, else None.

    "Tonight" is a real category rather than a flourish: a collection that
    starts at 20:00 today and one that starts at 08:00 today need different
    reactions, and "today" flattens the two.
    """
    when = _when(order)
    if not when:
        return None
    today = (today or datetime.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    days = (when.replace(hour=0, minute=0, second=0, microsecond=0) - today).days
    if days < 0 or days > URGENT_DAYS:
        return None
    if days == 0:
        return (0, "tonight" if when.hour >= 18 else "today")
    if days == 1:
        return (1, "tomorrow")
    return (days, f"in {days} days")


# ---------- the notice ----------
def build_notice(mapped, extract_name, today=None, rota_path=None):
    """The post-upload email, in the _pending_email.json shape.

    Returns (email, urgent) so the caller can print what was flagged; urgent is
    a list of dicts, one per order that needs another region to know about it.
    """
    by_first, me = team_emails()
    rota, people = {}, []
    path = rota_path or (SHEET_CACHE if os.path.exists(SHEET_CACHE) else "")
    if path:
        try:
            rota, people = parse_sheet(path)
        except Exception:
            rota, people = {}, []

    kinds, urgent = [], []
    for o in mapped:
        kind = material_of(o.get("Product / Description"))
        if kind and kind not in kinds:
            kinds.append(kind)
        u = urgency(o, today)
        if not u:
            continue
        try:
            import nr_regions
            region = nr_regions.region_of(o.get("D Postcode"))
        except Exception:
            region = ""
        days, phrase = u
        when = _when(o)
        row = "tippers" if kind in TIPPER_PRODUCTS else "flats"
        names = []
        if region and when:
            day = rota.get(region, {}).get(
                when.replace(hour=0, minute=0, second=0, microsecond=0), {})
            names = day.get(row) or day.get("flats") or day.get("tippers") or []
        # An order whose only cover is you needs no heads-up - you are the one
        # reading the extract. Skipping it keeps the section to what it claims
        # to be: other people's work. With no rota there is nothing to judge on,
        # so it stays in rather than being silently dropped.
        my_first = ""
        for first, addr in by_first.items():
            if addr.lower() == me:
                my_first = first
        if names and my_first and all(n.lower() == my_first for n in names):
            continue
        urgent.append({
            "ref": str(o.get("Customer Order No") or "").strip(),
            "region": region, "phrase": phrase, "days": days,
            "when": when.strftime("%d/%m/%Y %H:%M") if when else "",
            "product": kind or str(o.get("Product / Description") or "").strip(),
            "to_site": str(o.get("Delivery Point") or "").strip(),
            "to_pc": str(o.get("D Postcode") or "").strip(),
            "from_site": str(o.get("Site Name - Collection") or "").strip(),
            "names": names,
        })
    urgent.sort(key=lambda x: (x["days"], x["region"]))

    # Cc the materials team(s) whose product is actually in this upload. More
    # than one is normal - an extract is not single-product.
    cc, seen = [], set()
    for k in kinds:
        team = material_team(k)
        if team and team["email"] not in seen:
            seen.add(team["email"])
            cc.append(team["email"])

    # The planners are the To line and the materials team is Cc - that is the
    # way round the last two real sends went.
    #
    # WHO the planners are comes from the rota itself, not from the whole team
    # roster: the roster carries managers and people who never take a region,
    # and mailing all of them every upload is how a notice becomes noise. The
    # names on this week's sheet are exactly the people the upload affects.
    to = [by_first[n] for n in people if n in by_first and by_first[n].lower() != me]
    if not to:                                   # no rota - better a wide net than silence
        to = [e for e in by_first.values() if e.lower() != me]

    done = _summary(kinds)
    lines = [f"{done} done.", ""]
    if urgent:
        # Don't name the region number here - the active region lives in
        # config.json and a hardcoded "not Region 2" goes quietly wrong the
        # day this runs for anyone else.
        lines += ["Heads up - these are close and sit outside my region, so "
                  "they are yours rather than mine:", ""]
        for u in urgent:
            # Name the person if the rota knows them. When it does not, say so
            # plainly - an invented name on an urgent order is worse than none.
            who = " / ".join(u["names"]) if u["names"] else "cover not on the rota"
            lines.append(f"    {u['ref']}  ({u['region'] or 'region unknown'}) - {u['phrase']}")
            lines.append(f"        {u['product']}: {u['from_site']} -> "
                         f"{u['to_site']} {u['to_pc']}")
            lines.append(f"        collection {u['when']}   @{who}")
            lines.append("")
        lines.append("Shout if you want me to pick any of them up.")
    return {
        "to": "; ".join(to), "cc": "; ".join(cc), "name": "",
        "subject": f"FW: {extract_name}",
        "message": "\n".join(lines),
        "date": "", "area": "",
        "orders": [str(o.get("Customer Order No") or "").strip() for o in mapped],
        "product_codes": [], "materials": done, "site": "", "postcode": "",
        "source": f"Synergy upload {extract_name}", "attach": extract_name,
        # Marks this entry as OURS. date_query stages into the same file with a
        # source that also starts "Synergy upload", so the prefix cannot tell
        # the two apart and clearing on it would bin the date query.
        "kind": "upload_notice",
    }, urgent


def stage(mapped, extract_name, today=None, refresh=True):
    """Build the notice and put it in front of a human. Sends nothing."""
    if refresh:
        fetch_sheet()
    email, urgent = build_notice(mapped, extract_name, today=today)
    try:
        existing = json.load(open(PENDING, encoding="utf-8"))
        if not isinstance(existing, list):
            existing = []
    except Exception:
        existing = []
    # date_query may already have staged a query about this same upload and
    # both belong in the panel, so append rather than overwrite. But drop any
    # notice from an EARLIER upload first: date_query only rewrites the file
    # when it finds something, so without this a quiet run leaves yesterday's
    # notice sitting in Review & send looking like today's.
    existing = [e for e in existing if (e or {}).get("kind") != "upload_notice"]
    existing.append(email)
    with open(PENDING, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=1)
    return email, urgent


if __name__ == "__main__":
    import sys
    p = sys.argv[1] if len(sys.argv) > 1 else fetch_sheet()
    if not p:
        print("No Transport Planner sheet found (Outlook unreachable and no cache).")
        raise SystemExit(1)
    rota = parse_rota(p)
    print(f"Rota: {os.path.basename(p)}")
    for region in sorted(rota):
        for day in sorted(rota[region]):
            cover = rota[region][day]
            print(f"  {region}  {day:%a %d/%m}  "
                  f"flats={cover.get('flats')}  tippers={cover.get('tippers')}")
