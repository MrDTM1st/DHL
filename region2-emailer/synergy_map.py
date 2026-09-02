"""
Order-upload mapping: raw Synergy Haulier Extract -> enriched order rows (the
"Master Template with Mapping" logic) -> NR upload CSV (via nr_csv).

Replicates the Synergy Template File:
  * direct field copies from the extract
  * Supplier Details lookups by collection Site Name: Contact Name, Telephone,
    collection_time/_end (Collection Date + the site's loading-hour windows),
    Raised-by email (appended), collection Notes
  * postcode overrides for two sites; Shipment No derived from the Delivery
    Instructions when blank; Y/N fields default to N; a Serial of 0 -> blank
Collection sites not in the store are reported as UNMATCHED - the dashboard will
pop up to add them and remember (self-learning).

    python synergy_map.py "<raw extract .xlsx>"
"""
import os, json, sys, warnings
from datetime import datetime, timedelta
import openpyxl
import nr_csv, outbox

warnings.filterwarnings("ignore")
HERE = os.path.dirname(os.path.abspath(__file__))
STORE = json.load(open(os.path.join(HERE, "_synergy_sites.json"), encoding="utf-8"))
SITES = STORE.get("sites", {})
OVERRIDES = STORE.get("postcode_overrides", {})
# raw extract name -> an EXISTING Supplier Details site code. The pop-up used to
# offer only blank boxes, so pairing "Land Recovery Limited" with the site
# already in the sheet meant retyping its address, hours and phone by hand - and
# created a second near-duplicate site rather than resolving the first.
ALIASES = STORE.get("aliases", {})


def _akey(s):
    return str(s).strip().rstrip("-").strip().lower()


def add_alias(raw, code):
    """Remember that this extract name means an existing site. No new site is
    created, so the Supplier Details list stays the single source of truth."""
    raw, code = str(raw).strip(), str(code).strip()
    if not raw or not code:
        return ""
    STORE.setdefault("aliases", {})[_akey(raw)] = code
    with open(os.path.join(HERE, "_synergy_sites.json"), "w", encoding="utf-8") as f:
        json.dump(STORE, f, indent=1)
    globals()["ALIASES"] = STORE["aliases"]
    return code


def site_codes():
    """Every known collection site, for the dashboard dropdown."""
    return sorted(SITES, key=lambda s: s.lower())


def site_options():
    """[{code, pc, town}] - what the dropdown shows. The code alone is not
    enough to choose between sites that share an address."""
    out = []
    for c in sorted(SITES, key=lambda s: s.lower()):
        d = SITES.get(c) or {}
        out.append({"code": c,
                    "pc": str(d.get("postcode", "") or "").strip(),
                    "town": str(d.get("town", "") or "").strip()})
    return out


def _hours(s):
    try:
        h, m, sec = str(s).split(":")
        return timedelta(hours=int(h), minutes=int(m), seconds=int(float(sec)))
    except Exception:
        return None


def _as_dt(v):
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    if hasattr(v, "year"):                       # date
        return datetime(v.year, v.month, v.day)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(str(v).strip()[:10], fmt)
        except Exception:
            pass
    return None


def add_site(code, details):
    """Learn a new collection site: save it to the store so every future upload
    matches it automatically (the self-learning half of the site pop-up)."""
    global SITES
    code = str(code).strip()
    STORE.setdefault("sites", {})[code] = {
        "contact": details.get("contact", ""), "postcode": details.get("postcode", ""),
        "telephone": details.get("telephone", ""),
        "start_hours": details.get("start_hours", "") or "07:00:00",
        "close_hours": details.get("close_hours", "") or "17:00:00",
        "email": details.get("email", ""), "notes": details.get("notes", ""),
        # town rides along so the dashboard dropdown can tell four sites at one
        # postcode apart - ST6 4NU alone carries L13-Mossend LDC-, Land Recovery
        # Ltd and Land Recovery Ltd - ALSAGER, and the code names do not say which
        # is which.
        "town": details.get("town", ""), "address": details.get("address", ""),
    }
    with open(os.path.join(HERE, "_synergy_sites.json"), "w", encoding="utf-8") as f:
        json.dump(STORE, f, indent=1)
    SITES = STORE["sites"]
    return code


def map_orders(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def gi(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    C = dict(order=gi("customer order no"), ship=gi("shipment no"),
             site=gi("site name - collection"), ocn=gi("order contact name"),
             oco=gi("order contact no"), a1=gi("address1", "address 1"),
             a2=gi("address 2"), a3=gi("address 3"), pc=gi("postcode"),
             cdate=gi("collection date"), dpoint=gi("delivery point"),
             dcn=gi("d contact name"), da1=gi("d address1", "d address 1"),
             da2=gi("d address 2"), da3=gi("d address 3"), dpc=gi("d postcode"),
             dphone=gi("ship to contact phone", "d telephone no"),
             ddate=gi("delivery date"), dtime=gi("delivery_time"),
             dtimee=gi("delivery_time_end"), psc=gi("product / service code"),
             pd=gi("product / description"), qty=gi("product qty"),
             serial=gi("serial number"), instr=gi("shipping instructions", "delivery instructions"),
             raised=gi("raised by"), approver=gi("approver"), account=gi("account"),
             otype=gi("order type"), hiab=gi("hiab"), escort=gi("vehicle escort"),
             pts=gi("pts"), banksman=gi("banksman"), moffett=gi("moffett"),
             loggrab=gi("log grab"), rsteer=gi("rear steer"), vtype=gi("vehicle type"),
             cc=gi("cost centre"), ndel=gi("notes for delivery location comments"))

    def g(r, key):
        i = C.get(key)
        return r[i] if (i is not None and i < len(r) and r[i] is not None) else ""

    def yn(r, key):
        v = str(g(r, key)).strip()
        return v if v else "N"

    def nsite(s):   # tolerate trailing '-' / whitespace / case (extract vs store)
        return str(s).strip().rstrip("-").strip().lower()
    norm = {nsite(k): v for k, v in SITES.items()}

    from modules import site_matching
    deliv_store = site_matching.SiteStore(os.path.join(HERE, "_sites.json"))
    mapped, unmatched, held = [], {}, {}
    for r in rows[1:]:
        if not str(g(r, "order")).strip():
            continue
        site = str(g(r, "site")).strip()
        sd = SITES.get(site) or norm.get(nsite(site))
        if sd is None and site:                       # paired by hand previously?
            aliased = ALIASES.get(_akey(site))
            if aliased:
                sd = SITES.get(aliased) or norm.get(nsite(aliased))
        if sd is None and site:
            unmatched[site] = unmatched.get(site, 0) + 1
        sd = sd or {}

        # The EXACT match in the synergy flow is on the COLLECTION site only -
        # the template's Supplier Details VLOOKUP (Delali, 2026-07-20). The
        # Delivery Point passes through from the extract UNTOUCHED: never
        # canonicalised, never held - the template copies it verbatim, and a
        # held row here just delayed real orders. Brand-new delivery names are
        # still learned silently so the site list keeps growing.
        draw = str(g(r, "dpoint")).strip()
        dcanon = draw
        if draw:
            try:
                msite, res = deliv_store.match(draw)
                if not msite and not res:
                    deliv_store.add_sites([draw])
            except Exception:
                pass

        cdate = _as_dt(g(r, "cdate"))
        ct = cte = ""
        if cdate:
            sh, eh = _hours(sd.get("start_hours")), _hours(sd.get("close_hours"))
            if sh is not None:
                ct = (cdate + sh).strftime("%d/%m/%Y %H:%M")
            if eh is not None:
                cte = (cdate + eh).strftime("%d/%m/%Y %H:%M")
        # delivery times come straight from the extract (not derived)
        def dfmt(v):
            if isinstance(v, datetime):
                return v.strftime("%d/%m/%Y %H:%M")
            return str(v or "").strip()
        dt, dte = dfmt(g(r, "dtime")), dfmt(g(r, "dtimee"))

        ship = str(g(r, "ship")).strip()   # leave blank if the extract has none (no junk derivation)
        serial = g(r, "serial")
        serial = "" if str(serial).strip() in ("0", "") else serial
        raised, email = str(g(r, "raised")).strip(), sd.get("email", "")
        raisedby = (raised + ";" + email) if (raised and email) else (email or raised)

        mapped.append({
            "Customer Order No": g(r, "order"), "Shipment No": ship,
            "Site Name - Collection": site, "Contact Name": sd.get("contact", ""),
            "Address 1": g(r, "a1"), "Address 2": g(r, "a2"), "Address 3": g(r, "a3"),
            "Postcode": OVERRIDES.get(site, g(r, "pc")), "Telephone No": sd.get("telephone", ""),
            "collection time": ct, "collection time end": cte,
            "Delivery Point": dcanon, "D Contact Name": g(r, "dcn"),
            "D Address 1": g(r, "da1"), "D Address 2": g(r, "da2"), "D Address 3": g(r, "da3"),
            "D Postcode": g(r, "dpc"), "D Telephone No": g(r, "dphone"),
            "delivery time": dt, "delivery time end": dte,
            "Product / Service Code": g(r, "psc"), "Product / Description": g(r, "pd"),
            "Product Qty": g(r, "qty"), "Serial Number": serial,
            "Delivery Instructions": g(r, "instr"), "Raised by": raisedby,
            "Cost Centre": g(r, "cc") or None, "Account": g(r, "account"),
            "Vehicle Type": g(r, "vtype"), "HIAB": yn(r, "hiab"),
            "Vehicle Escort": yn(r, "escort"), "PTS": yn(r, "pts"),
            "Banksman": yn(r, "banksman"), "Moffett": yn(r, "moffett"),
            "Log Grab": yn(r, "loggrab"), "Rear Steer": yn(r, "rsteer"),
            "Notes for Collection Location Comments": sd.get("notes", ""),
            "Notes for Delivery Location Comments": g(r, "ndel"), "Est Cost": None,
        })
    return mapped, unmatched, held


UNMATCHED_FILE = os.path.join(HERE, "_synergy_unmatched.json")
NEWSITES_FILE = os.path.join(HERE, "_synergy_newsites.json")


TEMPLATE = os.path.join(HERE, "synergy_template.xlsx")


def _hkey(h):
    """Header comparison key: case/space/underscore-insensitive, so the raw
    extract's 'collection_time' matches the template's 'collection time'."""
    return "".join(ch for ch in str(h or "").lower() if ch.isalnum())


def fill_template(mapped, raw_path, out_path, held_sites=None):
    """Fill a COPY of the real Synergy Template File (the sheet the manual
    process 'goes through first to be formatted'):
      * raw extract rows pasted into 'Data Here' from row 4 (headers on row 3),
        aligned BY COLUMN NAME in case the extract's column order drifts;
      * our corrected mapped VALUES written over 'Master Template with Mapping'
        from row 3 - the template's own formulas reproduce old bugs (junk
        Shipment No, raw Heavy account), so the computed mapping wins;
      * Supplier Details + Vehicle Type sheets ride along untouched.
    Falls back to a plain one-sheet dump if the template file is missing."""
    if not os.path.exists(TEMPLATE):
        return _plain_sheet(mapped, out_path)
    import shutil, warnings
    warnings.filterwarnings("ignore")
    shutil.copy(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)
    dh, mt = wb["Data Here"], wb["Master Template with Mapping"]

    # --- Data Here: raw rows from row 4, aligned to the row-3 header names.
    # Rows with no order number and rows HELD for a site decision are skipped,
    # so Data Here stays 1:1 with the Master rows - the template's leftover
    # formula columns (Collection Date etc.) read their own row and must not
    # slip against our values.
    held_sites = {str(s).strip() for s in (held_sites or set())}
    raw = openpyxl.load_workbook(raw_path, data_only=True).active
    rrows = list(raw.iter_rows(values_only=True))
    rhdr = [_hkey(h) for h in rrows[0]]
    try:
        oi = rhdr.index(_hkey("Customer Order No"))
    except ValueError:
        oi = None
    try:
        di = rhdr.index(_hkey("Delivery Point"))
    except ValueError:
        di = None
    dcol = {}
    for c in range(1, dh.max_column + 1):
        k = _hkey(dh.cell(3, c).value)
        if k:
            dcol[k] = c
    out_r = 4
    for r in rrows[1:]:
        order = r[oi] if (oi is not None and oi < len(r)) else None
        if oi is not None and (order is None or not str(order).strip()):
            continue                      # same skip rule as map_orders
        if di is not None and di < len(r) and str(r[di] or "").strip() in held_sites:
            continue
        for i, v in enumerate(r):
            if i < len(rhdr) and rhdr[i] in dcol:
                dh.cell(out_r, dcol[rhdr[i]], v)
        out_r += 1

    # --- Master: our computed values over the formula rows (row 3 down)
    mcol = {}
    for c in range(1, mt.max_column + 1):
        k = _hkey(mt.cell(1, c).value)
        if k:
            mcol[k] = c
    for i, row in enumerate(mapped):
        r = 3 + i
        for key, val in row.items():
            c = mcol.get(_hkey(key))
            if c:
                mt.cell(r, c, val if val != "" else None)
    # clear leftover formula rows beyond our data so the sheet doesn't show
    # #N/A noise from formulas pointing at empty Data Here rows
    for r in range(3 + len(mapped), mt.max_row + 1):
        for c in range(1, mt.max_column + 1):
            if mt.cell(r, c).value is not None:
                mt.cell(r, c).value = None
    wb.save(out_path)
    return out_path


def _plain_sheet(mapped, path):
    """Fallback when the template file isn't present: a clean one-sheet dump."""
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synergy Upload"
    if mapped:
        cols = list(mapped[0].keys())
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="FFFF00")
        for row in mapped:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
    wb.save(path)
    return path


def _raised_by(mapped):
    """Who to ask about a bad date: whoever the extract says raised the orders,
    falling back to the team's own address rather than sending it nowhere."""
    for o in mapped:
        who = str(o.get("Raised by") or "").strip()
        if who:
            return who
    try:
        import build_drafts as bd
        return bd.team_config().get("me") or bd.DHL_SMTP
    except Exception:
        return ""


def main():
    args = sys.argv[1:]
    if args and args[0] == "addsites":
        # learn sites the dashboard collected (written to _synergy_newsites.json)
        try:
            sites = json.load(open(NEWSITES_FILE, encoding="utf-8"))
        except Exception:
            sites = {}
        learned, paired = [], []
        for code, det in sites.items():
            # {"pair_with": "<existing site code>"} means "this extract name IS
            # that site" - an alias, not a new entry.
            existing = str((det or {}).get("pair_with") or "").strip()
            if existing:
                add_alias(code, existing)
                paired.append(f"{code} -> {existing}")
            else:
                add_site(code, det)
                learned.append(code)
        if learned:
            print(f"Learned {len(learned)} new site(s): {', '.join(learned)}")
        if paired:
            print(f"Paired {len(paired)} name(s) to existing sites: {'; '.join(paired)}")
        if not learned and not paired:
            print("Nothing to learn.")
        return
    if not args or not os.path.exists(args[0]):
        print("Usage: python synergy_map.py <raw extract .xlsx>  |  synergy_map.py addsites"); return
    mapped, unmatched, held = map_orders(args[0])
    # record the misses for the dashboard site pop-up
    with open(UNMATCHED_FILE, "w", encoding="utf-8") as f:
        json.dump([{"site": s, "count": n} for s, n in sorted(unmatched.items(), key=lambda x: -x[1])],
                  f, indent=1)
    stamp = datetime.now().strftime("%d%m%Y%H%M%S")
    # the filled-in Synergy Template first (raw paste in Data Here + mapped
    # values in Master - the working record), THEN the NR upload CSV
    sheet = fill_template(mapped, args[0], outbox.path(f"Synergy Upload {stamp}.xlsx"),
                          held_sites=set(held))
    records = nr_csv.transform(mapped)
    out = nr_csv.write_csv(records, outbox.path(f"NR_upload_{stamp}.csv"))
    print(f"Mapped {len(mapped)} order line(s).")
    print(f"  SHEET: {sheet}")

    # A delivery timed before its own collection cannot be booked, and it is not
    # ours to correct - the extract's raiser has to say which end is wrong. The
    # email that asks is staged for Review & send; nothing goes on its own.
    import date_query
    backwards = date_query.raise_query(
        mapped, _raised_by(mapped), f"Synergy upload {stamp[:8]}")
    if backwards:
        print(f"  !! {len(backwards)} DELIVERY BEFORE COLLECTION - not bookable as they stand:")
        for p in backwards:
            print(f"       {p['ref']}  collect {p['collection']}  "
                  f"deliver {p['delivery']}  ({p['back_by']} earlier)")
        print("     An email asking them to confirm is ready for Review & send.")
    print(f"  BACKWARDS_DATES {len(backwards)}")
    if held:
        n = sum(held.values())
        print(f"  ~~ {n} row(s) HELD - delivery site needs a decision on the dashboard "
              f"(then re-run): {', '.join(sorted(held))}")
    if unmatched:
        print(f"  !! {len(unmatched)} collection site(s) NOT in the store - add these:")
        for s, n in sorted(unmatched.items(), key=lambda x: -x[1]):
            print(f"     {n:3}x  {s}")
    else:
        print("  all collection sites matched.")
    print(f"  CSV: {out}")


if __name__ == "__main__":
    main()
