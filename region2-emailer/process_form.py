"""
Process an already-filled Haulage Request Form (the usual ad hoc).

Reads the form's own "RHPC Admin - DHL USE ONLY" row - via Excel itself, so
the form's formulas produce the genuine values - and runs it through the
replicated database logic to build the NR upload CSV in the outbox.

    python process_form.py "<path to filled form.xlsx>"
    python process_form.py <reference or filename fragment>   # searches email
    python process_form.py latest                             # newest form in email
"""
import sys, os, re, json
from datetime import datetime
import nr_csv, outbox

HERE = os.path.dirname(os.path.abspath(__file__))
DHL_SMTP = "delali.opoku@dhl.com"
FORM_HINTS = ("haulage request", "transport request", "request form")


def find_form(query):
    import win32com.client
    ns = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    dhl = None
    for i in range(1, ns.Folders.Count + 1):
        if ns.Folders.Item(i).Name.lower() == DHL_SMTP:
            dhl = ns.Folders.Item(i)
            break

    def sub(f, name):
        if f is None:
            return None
        for i in range(1, f.Folders.Count + 1):
            c = f.Folders.Item(i)
            if c.Name.strip().lower() == name.strip().lower():
                return c

    inbox = sub(dhl, "Inbox")
    q = (query or "").lower()
    latest = q in ("", "latest")
    out = os.path.join(HERE, "_form.xlsx")
    for folder in (inbox, sub(sub(inbox, "ADHOC"), "DTS")):
        if folder is None:
            continue
        items = folder.Items
        try:
            items.Sort("[ReceivedTime]", True)
        except Exception:
            pass
        for it in items:   # full history, newest first
            try:
                subj = str(it.Subject or "").lower()
                for j in range(1, it.Attachments.Count + 1):
                    att = it.Attachments.Item(j)
                    fn = str(att.FileName)
                    if not fn.lower().endswith((".xlsx", ".xlsm")):
                        continue
                    fl = fn.lower()
                    hit = (any(h in fl or h in subj for h in FORM_HINTS)
                           if latest else (q in fl or q in subj))
                    if hit:
                        att.SaveAsFile(out)
                        return out, fn, folder.Name
            except Exception:
                continue
    return None, None, None


def read_rhpc_rows(path):
    """Open the form in Excel (invisible) so its formulas evaluate, and read
    the RHPC Admin rows. Falls back to cached values if Excel is unavailable."""
    rows = None
    try:
        import win32com.client
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.abspath(path), ReadOnly=True, UpdateLinks=0)
        try:
            names = [s.Name for s in wb.Worksheets]
            target = next((n for n in names if n.strip().lower().startswith("rhpc admin")), None)
            sh = wb.Worksheets(target)
            rows = sh.Range(sh.Cells(1, 1), sh.Cells(4, 60)).Value
        finally:
            wb.Close(False)
            xl.Quit()
    except Exception:
        import openpyxl, warnings
        warnings.filterwarnings("ignore")
        wbo = openpyxl.load_workbook(path, data_only=True)
        target = next((n for n in wbo.sheetnames if n.strip().lower().startswith("rhpc admin")), None)
        sho = wbo[target]
        rows = [[c.value for c in r] for r in sho.iter_rows(min_row=1, max_row=4, max_col=60)]
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[2:4]:   # data rows 3 and 4 (4 = return leg, if present)
        d = {headers[i]: r[i] for i in range(len(headers)) if headers[i]}
        if d.get("Customer Order No") not in (None, ""):
            out.append(d)
    return out


def fmt_dt(v):
    """Excel encodes 'no date' as the 1899/1900 epoch (a bare time cell reads
    as time(0,0), COM gives 1899-12-30) - those are placeholders for a date
    the requester hasn't set (e.g. a TBC return leg), NOT real dates. They
    must come out EMPTY, never '30/12/1899 00:00' in the upload."""
    import datetime as _dt
    if v is None or v == "":
        return ""
    if isinstance(v, _dt.datetime):
        return "" if v.year < 1990 else v.strftime("%d/%m/%Y %H:%M")
    if isinstance(v, _dt.time):
        return ""   # a bare time has no date - can't make an upload window
    try:
        return v.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(v)


ADHOC_ACCOUNT = "NRADHOC"   # default when the form leaves the account unset
UNSET_ACCOUNTS = {"", "please select", "select", "none"}


def _norm_order(ref):
    """Normalise an order number for the upload: spaces are not allowed, so a
    reference like 'FS-PLC CARDS' becomes 'FS-PLC-CARDS'. Collapses runs of
    whitespace/hyphens to a single hyphen and trims the ends. Excel hands
    numeric refs over as floats - 1770679.0 must upload as 1770679."""
    if isinstance(ref, float) and ref.is_integer():
        ref = int(ref)
    r = str(ref or "").strip()
    r = re.sub(r"\s+", "-", r)
    r = re.sub(r"-{2,}", "-", r).strip("-")
    return r


def _ref_incomplete(ref):
    """True only if the order number is missing or truncated - blank, or ending
    on a separator (e.g. a blank Collection Ref leaves a bare 'FS-'). A valid
    letters-only reference such as 'FS-PLC-CARDS' is NOT incomplete."""
    r = str(ref or "").strip()
    return (not r) or r[-1] in "-/ " or r.upper() in ("FS", "FS-")


def account_for(d):
    """Keep a preset account; only default to NRADHOC when the form left it
    on 'Please select' / blank."""
    acct = str(d.get("Account") or "").strip()
    return acct if acct.lower() not in UNSET_ACCOUNTS else ADHOC_ACCOUNT


def _dateonly_window(a, b):
    """True for Excel's 'date picked, NO time': a zero-length midnight window
    (start == end == 00:00 on the same day, or no end at all). A real
    midnight job always has a non-zero window (e.g. 00:00-02:00)."""
    import datetime as _dt
    mid = lambda v: (isinstance(v, _dt.datetime) and v.year >= 1990
                     and (v.hour, v.minute) == (0, 0))
    return mid(a) and (b in (None, "") or (mid(b) and b == a))


def to_transform_row(d):
    r = dict(d)
    r["collection time"] = fmt_dt(d.get("collection_time"))
    r["collection time end"] = fmt_dt(d.get("collection_time_end"))
    r["delivery time"] = fmt_dt(d.get("delivery_time"))
    r["delivery time end"] = fmt_dt(d.get("delivery_time_end"))
    # a date-only window (midnight-to-midnight) means the requester never set
    # a time - blank it so the 09:00-17:00 default below takes over
    if nr_csv.unstated_window(d.get("collection_time"), d.get("collection_time_end")):
        r["collection time"] = r["collection time end"] = ""
    if nr_csv.unstated_window(d.get("delivery_time"), d.get("delivery_time_end")):
        r["delivery time"] = r["delivery time end"] = ""
    # Delali (24/07): "if you ever get one where there is no delivery time,
    # just put nine to five so when I upload it the system recognizes it."
    # Only the WINDOW is invented - the date comes from the leg's own date
    # column (the other end's date as a last resort) and a form with no date
    # at all still comes out blank; a date is never made up.
    def _date_of(*keys):
        for k in keys:
            ds, _ = _dt_parts(d.get(k))
            if ds:
                return ds
        return ""
    # 09:01-17:01, not 09:00-17:00: the odd minute is how CTMS shows at a glance
    # that these times are still ours and nobody has confirmed them.
    if not r["collection time"]:
        base = _date_of("Collection Date", "collection_time", "Delivery Date", "delivery_time")
        if base:
            r["collection time"], r["collection time end"] = nr_csv.unconfirmed_window(base)
    if not r["delivery time"]:
        base = _date_of("Delivery Date", "delivery_time", "Collection Date", "collection_time")
        if base:
            # Default the delivery off THIS job's collection window (+1h01m),
            # which is what the real database does - not off a fixed pair.
            r["delivery time"], r["delivery time end"] = nr_csv.unconfirmed_window(
                base, (r["collection time"], r["collection time end"]))
    # A window with a start but no end used to be closed at the fixed 17:01,
    # which put the delivery marker on collection windows and turned a 22:00
    # start into an end earlier than itself. See nr_csv.close_window.
    for a, b in (("collection time", "collection time end"), ("delivery time", "delivery time end")):
        if r[a] and not r[b]:
            r[b] = nr_csv.close_window(r[a])
    r["Account"] = account_for(d)   # preset account wins; else NRADHOC
    return r


# ---------- dashboard map record ----------
# Each processed ad hoc is also saved as a record shaped like a tracker
# record, so the dashboard's brief, haulier ranking and map focus mode all
# work on it unchanged. The agent publishes these on the panel.
ADHOCS = os.path.join(HERE, "_adhocs.json")
FORMS_DIR = os.path.join(HERE, "_adhoc_forms")   # kept to forward with cover requests


def _flag(d, k):
    return str(d.get(k) or "N").strip().upper() not in ("N", "")


def _dt_parts(v):
    """(date_str, time_str) from an Excel cell that may hold a datetime, a
    bare time, or nothing. The 1899/1900 epoch and time(0,0) are Excel's
    'not set' placeholders (a TBC return leg) - they read as EMPTY."""
    import datetime as _dt
    if isinstance(v, _dt.datetime):
        if v.year < 1990:
            return "", (v.strftime("%H:%M") if (v.hour or v.minute) else "")
        return v.strftime("%d/%m/%Y"), v.strftime("%H:%M")
    if isinstance(v, _dt.time):
        return "", (v.strftime("%H:%M") if (v.hour or v.minute) else "")
    # Already-formatted strings, which is what the transform shape carries
    # (nr_csv.dts_row builds 'dd/mm/yyyy HH:MM'). These used to fall through to
    # ("", "") - so a DTS pinned on the map with no dates on it at all.
    for f in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
              "%Y-%m-%d %H:%M"):
        try:
            p = _dt.datetime.strptime(str(v).strip(), f)
        except Exception:
            continue
        if p.year < 1990:
            return "", (p.strftime("%H:%M") if (p.hour or p.minute) else "")
        return p.strftime("%d/%m/%Y"), p.strftime("%H:%M")
    return "", ""


def _covers(qty, prod):
    """True when the qty text already says what the product is ('12 x pallets'
    covers 'PALLET'), so the materials line doesn't need the product repeated.
    Compared on stemmed words: 'Cable Drums' is NOT covered by 'X1 DRUM'
    (the 'cable' is information) so that pair combines."""
    words = {w.rstrip("s") for w in re.findall(r"[a-z]+", qty.lower())}
    pw = [w.rstrip("s") for w in re.findall(r"[a-z]+", prod.lower())]
    return bool(pw) and all(w in words for w in pw)


def _leg_dates(d):
    """(collection_date, coll window, delivery_date, del window) for one row.
    The form has dedicated DATE columns, and collection/delivery can differ
    (collect Thu, deliver Sat) - never derive one date from the other. A
    date-only window (midnight-to-midnight) shows BLANK on the brief - the
    CSV gets the 9-5 default, but the brief never claims a time the
    requester didn't give."""
    # Two shapes reach this. The FORM shape has separate date columns and
    # underscore time keys; the TRANSFORM shape (nr_csv.dts_row, and anything
    # already through to_transform_row) has only the space-form keys, with the
    # date baked into them. Accept either, or a DTS pins on the map with no
    # dates at all.
    ctk = d.get("collection_time", d.get("collection time"))
    ctke = d.get("collection_time_end", d.get("collection time end"))
    dtk = d.get("delivery_time", d.get("delivery time"))
    dtke = d.get("delivery_time_end", d.get("delivery time end"))

    cdate, _ = _dt_parts(d.get("Collection Date"))
    cd2, ct1 = _dt_parts(ctk)
    _, ct2 = _dt_parts(ctke)
    ddate, _ = _dt_parts(d.get("Delivery Date"))
    dd2, dt1 = _dt_parts(dtk)
    _, dt2 = _dt_parts(dtke)
    if _dateonly_window(ctk, ctke):
        ct1 = ct2 = ""
    if _dateonly_window(dtk, dtke):
        dt1 = dt2 = ""
    return (cdate or cd2, {"earliest": ct1, "latest": ct2},
            ddate or dd2, {"earliest": dt1, "latest": dt2})


def _pair_return(rows):
    """The form's row 4 is the RETURN leg of row 3's order (same ref + '_R').
    One order, one email, one map record - so legs are paired, not split."""
    if len(rows) == 2:
        a = str(rows[0].get("Customer Order No") or "").strip()
        b = str(rows[1].get("Customer Order No") or "").strip()
        if a and b == a + "_R":
            return [(rows[0], rows[1])]
    return [(r, None) for r in rows]


def _adhoc_record(d, csv_name, ret=None, form_file="", kind="adhoc"):
    s = lambda k: str(d.get(k) or "").strip()

    cdate, cwin, ddate, dwin = _leg_dates(d)

    qty = d.get("Product Qty")
    if isinstance(qty, float) and qty.is_integer():
        qty = int(qty)
    qty = str(qty if qty is not None else "").strip()
    prod = s("Product / Description") or s("Product / Service Code")
    # quantities are allowed to be TEXT on these forms ("X1 DRUM",
    # "12 x pallets") - a plain number becomes an "Nx" multiplier, text that
    # already names the product stands alone, otherwise both are kept
    if re.fullmatch(r"\d{1,5}", qty):
        mats = qty + "x " + prod
    elif qty and prod and not _covers(qty, prod):
        mats = prod + " — " + qty
    else:
        mats = qty or prod
    # the weight/dimensions ride inside Delivery Instructions
    # ("... Qty X1 DRUM Weight - 400KG-DRUM W= 580MM X H 1000MM")
    wm = re.search(r"[Ww]eight\s*-?\s*(.+)", s("Delivery Instructions"))
    weight = wm.group(1).strip(" -") if wm else ""

    off = "HIAB" if _flag(d, "HIAB") else ("MOFFETT" if _flag(d, "Moffett") else "")
    rec = {
        "id": kind + "|" + s("Customer Order No") + "|" + datetime.now().strftime("%Y%m%d%H%M%S"),
        "kind": kind,
        "orders": [s("Customer Order No")],
        "site": s("Delivery Point"), "worksite": s("Delivery Point"),
        "postcode": s("D Postcode"),
        "collection_site": s("Site Name - Collection"), "collection_pc": s("Postcode"),
        "collections": [{"site": s("Site Name - Collection"), "pc": s("Postcode")}],
        "materials": mats,
        "qty": qty, "product": prod, "weight": weight,
        "delivery_date": ddate,
        "collection_date": cdate,
        "form_file": form_file,
        "details": {
            "time": dwin,
            "collection_time": cwin,
            "vehicle": {"value": s("Vehicle Type")},
            "offloading": {"value": off},
            "pts": {"value": "yes" if _flag(d, "PTS") else ""},
            "rear_steer": {"value": "yes" if _flag(d, "Rear Steer") else ""},
            "contact": {"name": s("D Contact Name"), "phone": s("D Telephone No")},
        },
        "csv": os.path.basename(csv_name),
        "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    if ret is not None:
        r2 = lambda k: str(ret.get(k) or "").strip()
        rcd, rcwin, rdd, rdwin = _leg_dates(ret)
        rec["return_leg"] = {
            "order": r2("Customer Order No"),
            "collection_date": rcd, "collection_time": rcwin,
            "delivery_date": rdd, "time": rdwin,
            "to_site": r2("Delivery Point"), "to_pc": r2("D Postcode"),
        }
    return rec


def save_adhocs(rows, csv_name, form_path="", keep=8, kind="adhoc"):
    """Newest first, capped - the map only ever needs the recent handful.
    Keeps a copy of the filled form so the cover-request email can forward it."""
    form_file = ""
    if form_path and os.path.exists(form_path) and rows:
        import shutil
        os.makedirs(FORMS_DIR, exist_ok=True)
        ref = re.sub(r"[^A-Za-z0-9._-]", "-", str(rows[0].get("Customer Order No") or "form"))
        form_file = "Haulage Request Form " + ref + os.path.splitext(form_path)[1]
        shutil.copyfile(form_path, os.path.join(FORMS_DIR, form_file))
    # An unreadable store must NOT become an empty one. This used to be
    # `except: old = []`, so a single failed read replaced the whole map with
    # just the job being processed - it silently binned five live records
    # twice in one day. Keep the bad file so it can be looked at, and never
    # start from empty while a non-empty file is sitting there.
    old = []
    if os.path.exists(ADHOCS):
        try:
            with open(ADHOCS, encoding="utf-8") as f:
                old = json.load(f)
            if not isinstance(old, list):
                raise ValueError("not a list")
        except Exception as ex:
            keep_name = ADHOCS + ".unreadable"
            try:
                os.replace(ADHOCS, keep_name)
                print(f"  !! _adhocs.json unreadable ({type(ex).__name__}) - kept as "
                      f"{os.path.basename(keep_name)}; the map starts from this job only")
            except Exception:
                pass
            old = []
    recs = [_adhoc_record(d, csv_name, ret=ret, form_file=form_file, kind=kind)
            for d, ret in _pair_return(rows)]
    # A job you have already booked off must not come back just because the form
    # was reprocessed. _booked_drops.json guarded the TRACKER only; nothing on
    # this path ever read it, so the same order was booked off three times in
    # fifteen minutes on 10/08 and a booked DTS reappeared twice.
    try:
        import tracker
        drops = tracker.booked_drops()
        kept = []
        for r in recs:
            if {str(o).strip() for o in r.get("orders", [])} & drops:
                print(f"SKIP: {'/'.join(r.get('orders', []))} is already booked in "
                      f"- not putting it back on the map")
            else:
                kept.append(r)
        recs = kept
    except Exception:
        pass
    # Same order reprocessed - replace its previous record rather than stacking
    # duplicate pins. Keyed on the ORDER NUMBERS, not the id: the id carries a
    # timestamp, so it is different on every run and never matched anything.
    # Reprocessing one form five times left five pins on the same spot, and the
    # stale ones kept whatever dates the form had at the time.
    fresh = {frozenset(str(o).strip() for o in r.get("orders", [])) for r in recs}
    old = [r for r in old
           if frozenset(str(o).strip() for o in r.get("orders", [])) not in fresh]
    tmp = ADHOCS + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump((recs + old)[:keep], f, indent=1)
    os.replace(tmp, ADHOCS)        # atomic: a concurrent reader never sees a partial file
    return len(recs)


def main():
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else "latest"
    if os.path.exists(arg):
        path, src = arg, "local file"
    else:
        print(f"Searching your mailbox for a filled form ({arg}, full history)...")
        path, fn, where = find_form(arg)
        if not path:
            print(f"NOT FOUND: no form matching '{arg}'.")
            return
        src = f"email attachment '{fn}' (folder: {where})"
    print(f"FORM: {src}")

    rows = read_rhpc_rows(path)
    if not rows:
        print("No data in the form's RHPC Admin row - is it actually filled in?")
        return

    # Order numbers can't contain spaces in the upload - hyphenate them (e.g.
    # 'FS-PLC CARDS' -> 'FS-PLC-CARDS') on both the order and shipment refs.
    for d in rows:
        d["Customer Order No"] = _norm_order(d.get("Customer Order No"))
        if str(d.get("Shipment No") or "").strip():
            d["Shipment No"] = _norm_order(d.get("Shipment No"))

    # a row 4 with the SAME ref as row 3 (no _R suffix) is a form-filling
    # duplicate, not a return leg - keeping it would book the order TWICE
    seen_refs, uniq = set(), []
    for d in rows:
        refn = str(d.get("Customer Order No"))
        if refn in seen_refs:
            print(f"NOTE {refn}: duplicate row on the form (same ref, not a return leg) - ignored.")
            continue
        seen_refs.add(refn)
        uniq.append(d)
    rows = uniq

    # Guard: a truncated order number (e.g. 'FS-' when the Collection Ref was
    # left blank) would be rejected by the upload. Refuse rather than produce a
    # dead file, and say exactly what to fix.
    good, bad = [], []
    for d in rows:
        (bad if _ref_incomplete(d.get("Customer Order No")) else good).append(d)
    for d in bad:
        print(f"!! INCOMPLETE ORDER NUMBER {str(d.get('Customer Order No')).strip()!r} - the form's "
              f"order-number field (e.g. Collection Ref) is blank. Fill it in and re-run; nothing written for this one.")
    if not good:
        print("Nothing written - no usable order number on the form.")
        return
    rows = good
    transformed = [to_transform_row(d) for d in rows]
    records = nr_csv.transform(transformed)

    # A delivery timed before its own collection is not bookable, and which end
    # is wrong is the requester's call, not ours. Checked on the TRANSFORMED
    # rows because that is where the date and time are finally combined - the
    # raw form columns keep them apart.
    import date_query
    backwards = date_query.raise_query(
        transformed,
        str(rows[0].get("Raised by") or "").split(";")[0].strip(),
        f"Ad hoc {rows[0].get('Customer Order No') or 'form'}")
    if backwards:
        print(f"!! {len(backwards)} DELIVERY BEFORE COLLECTION - not bookable as they stand:")
        for p in backwards:
            print(f"     {p['ref']}  collect {p['collection']}  "
                  f"deliver {p['delivery']}  ({p['back_by']} earlier)")
        print("   An email asking them to confirm is ready for Review & send.")
    print(f"BACKWARDS_DATES {len(backwards)}")
    # the order ref rides in the FILENAME - a Files card full of bare
    # timestamps gives no clue which CSV belongs to which job
    ref = re.sub(r"[^A-Za-z0-9]+", "-", str(rows[0].get("Customer Order No") or "")).strip("-")[:24]
    name = ("NR_heavy_" + (ref + "_" if ref else "")
            + datetime.now().strftime("%d%m%Y%H%M%S") + ".csv")
    out = nr_csv.write_csv(records, outbox.path(name))
    try:
        # Say what actually happened. save_adhocs can skip every record - a job
        # already booked off is refused, by design - and this printed "job
        # saved" regardless, so an upload that reached the map and one that was
        # silently refused produced identical output. AH19/8/26NESY was chased
        # for a quarter of an hour on the strength of that line.
        n_mapped = save_adhocs(rows, name, form_path=path)
        if n_mapped:
            print("MAP : job saved for the dashboard map (form kept for forwarding).")
        else:
            print("MAP : NOT on the map - every job in this form was refused, "
                  "see the SKIP line(s) above.")
    except Exception as ex:   # the map extra must never cost the CSV
        print(f"(map record not saved: {ex})")
    for d in rows:
        preset = str(d.get('Account') or '').strip().lower() not in UNSET_ACCOUNTS
        print(f"Order {d.get('Customer Order No')} | {d.get('Site Name - Collection')} "
              f"-> {d.get('Delivery Point')} | qty {d.get('Product Qty')} | "
              f"acct {account_for(d)}{' (preset)' if preset else ' (defaulted)'}")
        tr = to_transform_row(d)
        defaulted = []
        for lbl, rs, re_, k in (("collection", "collection_time", "collection_time_end", "collection time"),
                                ("delivery", "delivery_time", "delivery_time_end", "delivery time")):
            had_time = bool(fmt_dt(d.get(rs))) and not _dateonly_window(d.get(rs), d.get(re_))
            if not had_time and tr[k]:
                defaulted.append(lbl)
        missing = [lbl for lbl, k in (("collection", "collection time"),
                                      ("delivery", "delivery time")) if not tr[k]]
        if defaulted:
            print(f"NOTE {d.get('Customer Order No')}: {' & '.join(defaulted)} time not set on the "
                  f"form - defaulted to 09:00-17:00 in the CSV (correct in CTMS if wrong).")
        if missing:
            print(f"!! {d.get('Customer Order No')}: {' & '.join(missing)} has NO DATE on the form "
                  f"at all - times are BLANK in the CSV; fill in before uploading.")
    print(f"CSV : {out}")


if __name__ == "__main__":
    main()
