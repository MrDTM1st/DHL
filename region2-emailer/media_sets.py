"""
Media Sets: a week of courier runs -> the Network Rail upload CSV.

    python media_sets.py "Week 19 2026 Courier.xlsx"
    python media_sets.py "Week 19 2026 Courier.xlsx" 19      (force the week number)

The same shape as the Synergy upload, through a different form. The weekly
courier sheet carries 19 columns that line up one-for-one with the Media Sets
Input Sheet's columns B..T, and that form's RHPC Admin sheet is the SAME 47-
column contract the Synergy master sheet and the Haulage Request Form use - so
the finished rows go through nr_csv exactly like every other upload. Nothing
here is a second CSV writer.

Two things come out, both into the outbox:

  * the filled Media Sets Input Sheet - the human record, the thing you would
    have typed by hand, with the audit-check column intact
  * the NR upload CSV - what actually gets uploaded

Like synergy_map.py the CSV is built from values computed HERE, not read back
out of the saved workbook. The form's formulas are reproduced in code so the
output does not depend on Excel being installed, does not need COM, and cannot
blow the agent's 600-second subprocess timeout on a long week.

Sends nothing. Writes only into the outbox.
"""
import os
import re
import shutil
import sys
from datetime import datetime, date, time, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import nr_csv                      # noqa: E402  the ONE csv writer
import outbox                      # noqa: E402

TEMPLATE = os.path.join(HERE, "media_sets_template.xlsx")

# Baked into every row by the form itself (RHPC Admin, columns Y/Z/AF/AG/AP).
PRODUCT_CODE = "MEDIA_SETS"
# nr_csv's ITEMS line takes the DESCRIPTION first and only falls back to the
# code, so both carry MEDIA_SETS - otherwise the CSV would still read
# "Media Sets" however the code was spelled.
PRODUCT_DESC = "MEDIA_SETS"
ACCOUNT = "NRADHOC"
COST_CENTRE = "618294"
# A media set is always one item at 50kg, whatever the sheet happens to say -
# these are fixed for the product, not per-run figures somebody types in.
QTY = 1
WEIGHT_KG = 50
CONFIRMATION_EMAILS = ("Sophie.Robinson@networkrail.co.uk; "
                       "David.Whiston@networkrail.co.uk; "
                       "Richard.Wilkinson-Ford@networkrail.co.uk")

# The Input Sheet has 25 numbered rows but RHPC Admin only carries formulas for
# 23 of them (rows 3..25 <- input rows 5..27). Orders 24 and 25 would be typed
# in and then silently never reach the upload. We compute the CSV ourselves so
# OUR output is not capped - but the filled sheet still is, so say so.
SHEET_ROWS = 23

# The courier sheet repeats header names - "Address 1" is both D and L,
# "Postcode" is G and O, "Collection Time Start" is I and Q (the delivery pair
# is mislabelled in the source). Matching by name would mis-map those, so the
# contract is POSITION, and the headers are verified against it instead.
EXPECTED = [
    "site name - collection", "contact name", "order contact no",
    "address 1", "address 2", "address 3", "postcode",
    "collection date", "collection time start", "collection time end",
    "site name - delivery point", "address 1", "address 2", "address 3",
    "postcode", "delivery date", "collection time start", "collection time end",
    "product qty",
]


def _norm(h):
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for f in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip()[:10], f).date()
        except Exception:
            pass
    return None


def _as_time(v):
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, (int, float)):        # Excel serial fraction of a day
        secs = int(round(float(v) % 1 * 86400))
        return (datetime.min + timedelta(seconds=secs)).time()
    for f in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(str(v).strip(), f).time()
        except Exception:
            pass
    return None


def _stamp(d, t):
    """Date + time -> 'dd/mm/yyyy HH:MM', the way every other upload writes it.

    NOT a bare datetime. nr_csv stringifies whatever it is given, so a datetime
    object goes into the CSV as '2026-08-09 18:00:00' while the ad hoc and
    Synergy routes write '09/08/2026 18:00'. Two date formats in the same upload
    stream is exactly the kind of difference an import silently mishandles.
    """
    if d is None:
        return ""
    return datetime.combine(d, t or time(0, 0)).strftime("%d/%m/%Y %H:%M")


def week_of(path, sheet_name, override=None):
    """Week number: an explicit argument wins, then the sheet name, then the
    filename. 'Week 19' and 'Week 19 2026 Courier.xlsx' both give 19."""
    if override:
        return str(override).strip()
    for text in (sheet_name or "", os.path.basename(path)):
        m = re.search(r"week\s*(\d{1,2})", str(text), re.I)
        if m:
            return m.group(1)
    return ""


def read_courier(path):
    """(rows, sheet_name) - the 19 columns, positionally, headers verified."""
    import openpyxl
    import warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [_norm(ws.cell(row=1, column=c).value) for c in range(1, 20)]
    if hdr != EXPECTED:
        wb.close()
        diff = [f"col {chr(64 + i + 1)}: expected {EXPECTED[i]!r}, found {hdr[i]!r}"
                for i in range(min(len(hdr), len(EXPECTED))) if hdr[i] != EXPECTED[i]]
        raise SystemExit("The courier sheet's columns are not what Media Sets expects.\n  "
                         + "\n  ".join(diff or ["column count differs"])
                         + "\nNothing written - the columns repeat names, so guessing "
                           "which is which would silently swap collection and delivery.")
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
        if all(v in (None, "") for v in vals):
            continue
        if not str(vals[0] or "").strip():        # no collection site = not a job
            continue
        rows.append(vals)
    name = ws.title
    wb.close()
    return rows, name


def to_orders(rows, week):
    """The 47-column dicts, reproducing the form's RHPC Admin formulas."""
    out, problems = [], []
    for i, v in enumerate(rows, start=1):
        (csite, contact, phone, a1, a2, a3, pc, cdate, cstart, cend,
         dsite, d1, d2, d3, dpc, ddate, dstart, dend, _sheet_qty) = v

        cd, dd = _as_date(cdate), _as_date(ddate)
        ct = _stamp(cd, _as_time(cstart))
        cte = _stamp(cd, _as_time(cend))
        dt = _stamp(dd, _as_time(dstart))
        dte = _stamp(dd, _as_time(dend))

        # The form's own audit check: collection must not be after delivery.
        if isinstance(ct, datetime) and isinstance(dt, datetime) and ct > dt:
            problems.append(f"row {i} ({csite} -> {dsite}): collection "
                            f"{ct:%d/%m %H:%M} is AFTER delivery {dt:%d/%m %H:%M}")

        # "MS"&week&"/"&n&"/"&DAY&"/"&MONTH&"/"&YEAR&"/"&LEFT(postcode,2)
        ref = ""
        if cd:
            ref = (f"MS{week}/{i}/{cd.day}/{cd.month}/{cd.year}/"
                   f"{str(pc or '').strip()[:2].upper()}")

        out.append({
            "Customer Order No": ref, "Shipment No": ref,
            "Site Name - Collection": csite,
            "Contact Name": contact, "Order Contact Name": contact,
            "Order Contact No": phone, "Telephone No": phone,
            "Address 1": a1, "Address 2": a2, "Address 3": a3, "Postcode": pc,
            "Collection Date": cd,
            "collection time": ct, "collection time end": cte,
            "Delivery Point": dsite, "D Contact Name": contact,
            "D Address 1": d1, "D Address 2": d2, "D Address 3": d3,
            "D Postcode": dpc, "D Telephone No": phone,
            "Delivery Date": dd,
            "delivery time": dt, "delivery time end": dte,
            "Product / Service Code": PRODUCT_CODE,
            "Product / Description": PRODUCT_DESC,
            "Product Qty": QTY,
            # The NR upload has no weight column (nr_csv:182) - on the ad hoc
            # route the weight rides in the delivery instructions, so it does
            # the same here rather than being dropped on the floor.
            "Delivery Instructions": f"{PRODUCT_DESC} Qty {QTY} Weight {WEIGHT_KG}kg",
            "Raised by": CONFIRMATION_EMAILS,
            "Account": ACCOUNT, "Order Type": "C",
            "Vehicle Escort": "N", "PTS": "N", "Banksman": "N", "Log Grab": "N",
            "Cost Centre": COST_CENTRE,
        })
    return out, problems


def fill_sheet(rows, week, out_path):
    """A copy of the real Input Sheet with this week's rows typed in."""
    if not os.path.exists(TEMPLATE):
        print(f"  (no {os.path.basename(TEMPLATE)} here - skipping the filled sheet)")
        return None
    import openpyxl
    import warnings
    warnings.filterwarnings("ignore")
    shutil.copyfile(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)          # keep the formulas
    ws = wb["Input Sheet"]
    if week:
        ws["C1"] = int(week) if str(week).isdigit() else week
    for n, v in enumerate(rows[:SHEET_ROWS]):
        r = 5 + n
        for c, val in enumerate(v):                # data A..S -> sheet B..T
            ws.cell(row=r, column=2 + c).value = val
    wb.save(out_path)
    wb.close()
    return out_path


# Parcel Pass carry the media sets - small, single-item runs. The address is
# the one already used for them in Sent Items, not the personal ones.
PARCEL_PASS = "NR@Passlogistics.co.uk"


def cover_request(orders, week, sheet_path):
    """The 'can you cover this week?' email, staged for Review & send.

    The postcodes listed are the COLLECTIONS. Every media set delivers to the
    same place - EMCC Derby - so a list of delivery postcodes is one line and
    tells a courier nothing. Where the driver actually travels is the pickups,
    and that is what decides whether they can cover.
    """
    seen = {}
    for o in orders:
        pc = str(o.get("Postcode") or "").strip()
        if not pc:
            continue
        seen.setdefault(pc, {"n": 0, "site": str(o.get("Site Name - Collection") or "")})
        seen[pc]["n"] += 1
    dests = sorted({str(o.get("D Postcode") or "").strip()
                    for o in orders if str(o.get("D Postcode") or "").strip()})
    dates = sorted({o["Collection Date"] for o in orders if o.get("Collection Date")})
    when = (f"{dates[0]:%d/%m} to {dates[-1]:%d/%m}" if len(dates) > 1
            else (f"{dates[0]:%d/%m}" if dates else ""))

    lines = ["Hi,", "",
             f"Please see attached the media sets for week {week}"
             + (f", {when}" if when else "") + ".",
             "",
             f"Would you be able to cover? There {'are' if len(orders) != 1 else 'is'} "
             f"{len(orders)} run{'s' if len(orders) != 1 else ''}, collecting from the "
             f"postcodes below and delivering to "
             f"{', '.join(dests) if dests else 'the address on the sheet'}.",
             ""]
    for pc, d in sorted(seen.items()):
        lines.append(f"    {pc:10} {d['site']}" + (f"  (x{d['n']})" if d["n"] > 1 else ""))
    lines += ["", "Dates and times for each run are on the attached sheet.",
              "", "Let me know if you can take them and I will get them raised."]

    return {
        "to": PARCEL_PASS, "cc": "", "name": "",
        "subject": f"Media sets week {week} - can you cover?",
        "message": "\n".join(lines),
        "date": f"{dates[0]:%d/%m/%Y}" if dates else "",
        "area": "", "orders": [o["Customer Order No"] for o in orders if o.get("Customer Order No")],
        "product_codes": [PRODUCT_CODE], "materials": f"{len(orders)}x media sets",
        "site": "", "postcode": "", "source": f"media sets week {week}",
        "attach": [sheet_path] if sheet_path else [],
        # a haulier cover request, NOT a delivery contact - see send_emails
        "no_track": True, "haulier": "PARCELPASS", "_metric": "haulier_request",
    }


def _only(args):
    """Row numbers from --only 1  or  --only 1,4,9. The numbers are the # on the
    sheet, which is what the order reference carries."""
    for i, a in enumerate(args):
        if a.lower() in ("--only", "--rows") and i + 1 < len(args):
            return {int(x) for x in re.split(r"[\s,]+", args[i + 1]) if x.strip().isdigit()}
        if a.lower().startswith("--only="):
            return {int(x) for x in re.split(r"[\s,]+", a.split("=", 1)[1]) if x.strip().isdigit()}
    return set()


def main():
    argv = [a for a in sys.argv[1:]]
    only = _only(argv)
    args = []
    skip = False
    for i, a in enumerate(argv):                      # strip the flag + its value
        if skip:
            skip = False
            continue
        if a.lower() in ("--only", "--rows"):
            skip = True
            continue
        if a.lower().startswith("--only="):
            continue
        args.append(a)
    if not args or not os.path.exists(args[0]):
        print(__doc__)
        print("MEDIA_RESULT orders=0")
        return 2
    src = args[0]
    override = args[1] if len(args) > 1 else None

    rows, sheet_name = read_courier(src)
    week = week_of(src, sheet_name, override)
    print(f"Media Sets - week {week or '?'} - {len(rows)} courier run(s) "
          f"from {os.path.basename(src)}\n")
    if not rows:
        print("Nothing to do - no rows with a collection site.")
        print("MEDIA_RESULT orders=0")
        return 0
    if not week:
        print("  !! no week number in the sheet name or filename - the order")
        print("     references will read 'MS//...'. Pass one: media_sets.py <file> 19")

    orders, problems = to_orders(rows, week)

    # --only re-issues single runs, for when one order does not take in CTMS and
    # the rest did. The orders are numbered from the WHOLE sheet first and only
    # then filtered, so a re-issued row keeps the reference it was given the
    # first time - renumbering a filtered set would produce MS19/1 for what CTMS
    # already knows as MS19/7, and land a second, differently-named order.
    if only:
        picked = [o for n, o in enumerate(orders, start=1) if n in only]
        missing = sorted(only - set(range(1, len(orders) + 1)))
        if missing:
            print(f"  !! no row {', '.join(str(m) for m in missing)} on this sheet "
                  f"- it has {len(orders)}.")
        if not picked:
            print("Nothing selected - nothing written.")
            print("MEDIA_RESULT orders=0")
            return 1
        print(f"  RE-ISSUE: row(s) {', '.join(str(n) for n in sorted(only))} only, "
              f"of {len(orders)} on the sheet\n")
        orders = picked

    for o in orders:
        print(f"  {o['Customer Order No'] or '(no ref)':26} "
              f"{str(o['Site Name - Collection'])[:26]:26} {str(o['Postcode'] or ''):9}"
              f" -> {str(o['Delivery Point'])[:12]:12} {str(o['D Postcode']):9}")

    # An unrecognised outward code ships as the literal 'UNKNOWN' rather than
    # failing, so count them here instead of letting them through unremarked.
    unknown = [o for o in orders if nr_csv.region_of(o.get("Postcode")) == "UNKNOWN"
               or nr_csv.region_of(o.get("D Postcode")) == "UNKNOWN"]
    if unknown:
        print(f"\n  !! {len(unknown)} row(s) have a postcode with no region - they will "
              f"upload as UNKNOWN:")
        for o in unknown[:8]:
            print(f"       {o['Customer Order No']}  {o['Postcode']} -> {o['D Postcode']}")
    # Dates that run backwards are a question for whoever raised the sheet, not
    # something to guess at here - so the email that asks is staged for review
    # rather than the run quietly carrying on.
    import date_query
    backwards = date_query.raise_query(
        orders, CONFIRMATION_EMAILS, f"Media sets week {week or '?'}")
    if backwards:
        print(f"\n  !! {len(backwards)} DELIVERY BEFORE COLLECTION - not bookable as they stand:")
        for p in backwards:
            print(f"       {p['ref']}  collect {p['collection']}  "
                  f"deliver {p['delivery']}  ({p['back_by']} earlier)")
        print("     An email asking them to confirm is ready for Review & send.")
    if problems and not backwards:
        print(f"\n  !! {len(problems)} DATE ERROR(S) - the form would flag these:")
        for p in problems:
            print(f"       {p}")
    if len(rows) > SHEET_ROWS:
        print(f"\n  !! {len(rows)} rows but the Input Sheet's RHPC Admin only carries "
              f"{SHEET_ROWS} - the filled sheet shows the first {SHEET_ROWS}. "
              f"The CSV below has all {len(rows)}.")

    stamp = datetime.now().strftime("%d%m%Y%H%M%S")
    if only:
        # A re-issue is the CSV only. The filled sheet is the record of the
        # whole week and already exists; a one-row copy of it would just be
        # another thing to mistake for the real one.
        sheet_out = None
        tag = "-".join(str(n) for n in sorted(only))
        csv_out = outbox.path(f"NR_media_sets_wk{week or 'x'}_row{tag}_{stamp}.csv")
    else:
        sheet_out = fill_sheet(rows, week, outbox.path(f"Media Sets Week {week or 'x'} {stamp}.xlsx"))
        csv_out = outbox.path(f"NR_media_sets_wk{week or 'x'}_{stamp}.csv")
    nr_csv.write_csv(nr_csv.transform(orders), csv_out)

    print()
    if sheet_out:
        print(f"  SHEET: {sheet_out}")
    print(f"  CSV  : {csv_out}")

    # A clean week can go straight to the courier. A week with backwards dates
    # cannot - the date query is already staged in _pending_email.json and must
    # not be overwritten by a cover request for runs nobody has confirmed yet.
    # A re-issue must NOT stage a cover request: Parcel Pass were already asked
    # about this week, and a second email listing one postcode reads like a new
    # job rather than a fix.
    if not backwards and not only:
        import json as _json
        email = cover_request(orders, week or "?", sheet_out)
        with open(os.path.join(HERE, "_pending_email.json"), "w", encoding="utf-8") as f:
            _json.dump([email], f, indent=1)
        pcs = len({str(o.get("Postcode") or "").strip() for o in orders
                   if str(o.get("Postcode") or "").strip()})
        print(f"\n  COVER: email to {PARCEL_PASS} ready for Review & send "
              f"- {len(orders)} run(s) from {pcs} collection postcode(s), sheet attached.")
    print(f"  COVER_READY {1 if (not backwards and not only) else 0}")
    try:
        import metrics
        metrics.log("media_sets_built", orders=[o["Customer Order No"] for o in orders],
                    what=f"week {week}")
    except Exception:
        pass
    print(f"MEDIA_RESULT orders={len(orders)} backwards={len(backwards)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
