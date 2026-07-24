"""Haulier directory - who they are, what they can do, where they are.

Imports the "Haulier Contact List - Planner Version" workbook into
_hauliers.json (gitignored - contacts), then answers the planner's real
question: WHO SHOULD I RING FOR THIS JOB?

    python hauliers.py import "<contact list .xlsx>"
    python hauliers.py find --pc DN16 --need "Rail / S&C" "Artic Hiab" [--pts]
    python hauliers.py show "Lawsons"

Recommendation = capability match (hard filter) then ranked by distance from
the collection postcode, tier, and any quote history for that lane
(quotes.py). Distance is straight-line from postcode centroids - good enough
to rank "who's nearest", not a routing engine.
"""
import os, re, sys, json, math
import postcodes

HERE = os.path.dirname(os.path.abspath(__file__))
PATH = os.path.join(HERE, "_hauliers.json")

# capability columns start here in the Hauliers sheet (after the contact block)
_FIRST_CAP = "Bags"
_CONTACT_COLS = {"haulier name", "updated", "location", "postcode", "allocation",
                 "daytime phone", "ooh phone", "email contacts"}
_META_COLS = {"fors status & id", "oracle id", "cfx status"}


def _yes(v):
    return str(v or "").strip().upper().startswith("Y")


def _clean(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _emails(v):
    """The sheet packs several addresses into one cell, '/'-separated."""
    return [e.strip() for e in re.split(r"[/;,]", str(v or "")) if "@" in e]


# postcodes.py is the single correct implementation - the old pattern here read
# "DN3 1ED" as outcode DN31 (Grimsby, ~40mi from Doncaster), so the fallback
# used when a full postcode won't geocode landed the haulier in the wrong county.
_outward = postcodes.outward


# ---- coverage overrides ------------------------------------------------
# Operational facts from Delali that the contact list doesn't record. Applied
# in load() (not baked into the store) so re-importing a newer contact sheet
# never wipes them. Keys match by substring of the haulier name.
#
# Postcode AREAS considered "the north" for coverage purposes - Yorkshire/
# Humberside, the North East, the North West and Scotland. The Midlands belt
# (B CV DY WS WV ST TF LE DE NG LN NN...) is deliberately NOT in here.
NORTH_AREAS = sorted({
    # Yorkshire & Humberside
    "S", "DN", "HU", "YO", "LS", "BD", "HX", "HD", "WF", "HG",
    # North East
    "NE", "SR", "DH", "TS", "DL",
    # North West
    "CA", "LA", "PR", "FY", "BB", "BL", "OL", "M", "SK", "WA", "WN", "L", "CH",
    # Scotland
    "AB", "DD", "DG", "EH", "FK", "G", "HS", "IV", "KA", "KW", "KY", "ML",
    "PA", "PH", "TD", "ZE",
})

_OVERRIDES = {
    # "HHL mostly focus on the Midlands and southern/London region, so they
    # wouldn't really be doing northern work... HHL is only for DELIVERY, not
    # both" - Delali, 2026-07-22. Collecting FROM the north is fine (loads
    # originate at the northern steelworks and flow south all the time); what
    # they don't do is DELIVER up north. So the no-go only checks the
    # delivery end.
    "hotspur": {"no_go_areas": NORTH_AREAS, "no_go_scope": "delivery"},
    # "DHL NOC doesn't like doing night jobs and weekend jobs" - Delali,
    # 2026-07-24. Not a hard no: they stay on the list, but a night/weekend
    # job flags them and drops them below the external hauliers so the first
    # call is someone who'll actually take it.
    "dhl": {"avoid_night_weekend": True},
}


# Parcel Pass - the parcel/pallet NETWORK the desk books small ad hoc loads
# through (boxes/parcels/pallets on a transit van / 7.5t / 18t, no lifting
# kit). Not on the contact sheet and not a ranked haulier: it has no depot
# postcode and never joins the ring-round - the dashboard's ad hoc brief
# offers it directly. Lives in code (like _OVERRIDES) so a contact-list
# re-import never wipes it.
PARCEL_PASS = {
    "name": "Parcel Pass", "location": "", "postcode": "", "outward": "",
    "allocation": "", "tier": "", "do_not_use": False, "own_fleet": False,
    "phone": "0330 122 8447", "ooh": "", "emails": ["NR@Passlogistics.co.uk"],
    "caps": [], "ctms": "", "fors": "", "cfx": "", "parcel_service": True,
}

# The Network Rail MATERIALS TEAMS - the escalation route when an order's
# site contact can't be reached after a few tries: ask the team that owns
# the material for an alternative contact (Delali, 24/07). Every address
# verified against the team's real emails in the mailbox, never guessed
# (yes, the sleepers address genuinely contains an ampersand).
MATERIALS_TEAMS = {
    "rails":    {"name": "Steel Materials", "email": "SteelMaterials@networkrail.co.uk"},
    "ballast":  {"name": "Track Aggregates", "email": "Track.Aggregates@networkrail.co.uk"},
    "sleepers": {"name": "SCO Sleepers & Troughing", "email": "SCOSleepers&Troughing@networkrail.co.uk"},
}


def _apply_overrides(d):
    for h in d.get("hauliers", []) + d.get("couriers", []):
        nm = str(h.get("name", "")).lower()
        for key, ov in _OVERRIDES.items():
            if key in nm:
                h.update(ov)
    d["services"] = [dict(PARCEL_PASS)]
    return d


def load():
    try:
        return _apply_overrides(json.load(open(PATH, encoding="utf-8")))
    except Exception:
        return _apply_overrides({"hauliers": [], "couriers": [], "ctms": {}})


def save(d):
    tmp = PATH + ".tmp"
    json.dump(d, open(tmp, "w", encoding="utf-8"), indent=1)
    os.replace(tmp, PATH)


def _fill(cell):
    """The name cell's fill - the sheet encodes tier THERE, not in text."""
    try:
        f = cell.fill
        if f and f.fgColor:
            if f.fgColor.type == "rgb":
                return str(f.fgColor.rgb)
            if f.fgColor.type == "theme":
                return f"theme{f.fgColor.theme}"
    except Exception:
        pass
    return ""


def _tier_from_fill(fill):
    """Per the sheet's own 'Haulier Key' legend:
         blue (theme4)  = Tier 1 / Fleet
         no fill        = Tier 2
         RED (FFFF0000) = DO NOT USE  <- must never be recommended
    """
    if fill.startswith("FFFF0000") or fill == "FFFF0000":
        return "do_not_use"
    if fill.startswith("theme4"):
        return "tier1"
    return "tier2"


def _is_own_fleet(name):
    """DHL's OWN fleet (DHL NOC) - our vehicles, so it's approached before any
    external haulier. The sheet lumps it into the blue 'Tier 1 / Fleet' band,
    but Delali's order of approach is: own fleet -> tier 1 -> tier 2."""
    return "dhl" in str(name or "").lower()


def rank_of(h):
    """0 = our own fleet, 1 = tier 1, 2 = tier 2. Distance only ever breaks
    ties WITHIN a band - a nearer tier-2 never jumps a tier-1."""
    if h.get("own_fleet"):
        return 0
    return 1 if h.get("tier") == "tier1" else 2


# legend/section rows that live in the same column as the names
_NOT_A_HAULIER = ("haulier key", "courier key", "tier 1", "tier 2", "do not use",
                  "region allocation", "s1 -", "s2 -", "s3 -", "s4 -")


def import_workbook(path):
    """Parse the contact list into the store. Re-runnable - a newer version of
    the sheet just replaces it. Reads the name-cell FILL for tier/do-not-use."""
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)      # values
    wbf = openpyxl.load_workbook(path)                     # styles (fills)

    def sheet(*names):
        for n in wb.sheetnames:
            if n.strip().lower() in [x.lower() for x in names]:
                return wb[n]
        return None

    out = {"hauliers": [], "couriers": [], "ctms": {}}

    ct = sheet("CTMS Names")
    if ct:
        for r in ct.iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]:
                out["ctms"][_clean(r[0]).lower()] = _clean(r[1])

    for key, sh, capstart in (("hauliers", sheet("Hauliers"), _FIRST_CAP),
                              ("couriers", sheet("Couriers"), "Small Van")):
        if sh is None:
            continue
        shf = wbf[sh.title]
        hdr = [_clean(sh.cell(1, c).value) for c in range(1, sh.max_column + 1)]
        low = [h.lower() for h in hdr]
        try:
            ci = low.index(capstart.lower())
        except ValueError:
            ci = len(hdr)
        for ri, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
            name = _clean(row[0] if row else "")
            if not name or name.lower().startswith(_NOT_A_HAULIER):
                continue                     # legend / section row, not a haulier
            def g(col):
                try:
                    i = low.index(col)
                    return row[i] if i < len(row) else ""
                except ValueError:
                    return ""
            caps = [hdr[i] for i in range(ci, min(len(hdr), len(row)))
                    if hdr[i] and hdr[i].lower() not in _META_COLS and _yes(row[i])]
            tier = _tier_from_fill(_fill(shf.cell(ri, 1)))
            out[key].append({
                "name": name, "location": _clean(g("location")),
                "postcode": _clean(g("postcode")).upper(),
                "outward": _outward(g("postcode")),
                "allocation": _clean(g("allocation")),
                "tier": tier, "do_not_use": tier == "do_not_use",
                "own_fleet": _is_own_fleet(name) and tier != "do_not_use",
                "phone": _clean(g("daytime phone")), "ooh": _clean(g("ooh phone")),
                "emails": _emails(g("email contacts")),
                "caps": caps,
                "ctms": out["ctms"].get(name.lower(), ""),
                "fors": _clean(g("fors status & id")), "cfx": _clean(g("cfx status")),
            })
    save(out)
    return out


# ---- geography ---------------------------------------------------------
def _haversine(a, b):
    (la1, lo1), (la2, lo2) = a, b
    r = 6371.0
    p1, p2 = math.radians(la1), math.radians(la2)
    dp, dl = math.radians(la2 - la1), math.radians(lo2 - lo1)
    h = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(h)) * 0.621371     # miles


def geo_cache():
    try:
        return json.load(open(os.path.join(HERE, "_pc_geo.json"), encoding="utf-8"))
    except Exception:
        return {}


_norm_pc = postcodes.norm      # 'LE12 9 BS' / 'dn161bp' -> 'LE12 9BS' / 'DN16 1BP'


def geocode(postcodes):
    """Postcode centroids via postcodes.io, cached on disk. Industrial sites
    often have TERMINATED postcodes that return nothing (DN16 1BP, the British
    Steel depot) - those fall back to the OUTCODE centroid, which is plenty
    accurate for ranking who's nearest."""
    import urllib.request
    cache = geo_cache()
    want = {_norm_pc(p) for p in postcodes if p and _norm_pc(p)}
    need = sorted(w for w in want if cache.get(w) is None and w not in cache)
    for i in range(0, len(need), 90):
        try:
            req = urllib.request.Request(
                "https://api.postcodes.io/postcodes",
                data=json.dumps({"postcodes": need[i:i + 90]}).encode(),
                headers={"Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=20) as r:
                res = json.loads(r.read()).get("result") or []
            for x in res:
                q = _norm_pc(x.get("query", ""))
                rr = x.get("result")
                cache[q] = [rr["latitude"], rr["longitude"]] if rr else None
        except Exception:
            break
    # outcode fallback for anything still unresolved
    for w in sorted(w for w in want if not cache.get(w)):
        oc = _outward(w)
        if not oc:
            continue
        key = f"OUTCODE:{oc}"
        if key not in cache:
            try:
                with urllib.request.urlopen(
                        f"https://api.postcodes.io/outcodes/{oc}", timeout=15) as r:
                    rr = json.loads(r.read()).get("result") or {}
                cache[key] = [rr["latitude"], rr["longitude"]] if rr.get("latitude") else None
            except Exception:
                cache[key] = None
        if cache.get(key):
            cache[w] = cache[key]
    json.dump(cache, open(os.path.join(HERE, "_pc_geo.json"), "w", encoding="utf-8"))
    return cache


# ---- the actual question ----------------------------------------------
def recommend(from_pc, needs=(), to_pc="", limit=None, include_couriers=False):
    """Who should I ring for this job? Returns EVERYONE who fits (Delali wants
    the full list, not a top few), fleet -> tier 1 -> tier 2, closest to
    furthest within each band. Capability is a HARD filter (no point ranking
    someone who can't carry it), and so is coverage - a haulier is never
    suggested for a job in an area they don't work (HHL never gets northern
    jobs)."""
    d = load()
    pool = list(d.get("hauliers", []))
    if include_couriers:
        pool += list(d.get("couriers", []))
    needs = [str(n).strip().lower() for n in needs if str(n).strip()]
    c_area, d_area = postcodes.area(from_pc), postcodes.area(to_pc)

    def outside_coverage(h):
        """no_go_scope says WHICH end the no-go areas apply to: 'delivery'
        (HHL - happy to collect up north, won't deliver there), 'collection',
        or 'both' (the default)."""
        areas = set(h.get("no_go_areas", []))
        if not areas:
            return False
        scope = h.get("no_go_scope", "both")
        if scope in ("both", "collection") and c_area and c_area in areas:
            return True
        if scope in ("both", "delivery") and d_area and d_area in areas:
            return True
        return False

    ok = []
    for h in pool:
        if h.get("do_not_use"):
            continue                     # marked DO NOT USE on the sheet - never suggest
        if outside_coverage(h):
            continue
        caps = [c.lower() for c in h.get("caps", [])]
        if all(any(n in c for c in caps) for n in needs):
            ok.append(h)
    cache = geocode([from_pc] + [h.get("postcode") for h in ok])
    origin = cache.get(_norm_pc(from_pc))
    try:
        import quotes
        est = quotes.estimate(from_pc, to_pc) if to_pc else None
    except Exception:
        est = None
    known = {q["haulier"].lower() for q in (est or {}).get("quotes", [])}
    for h in ok:
        g = cache.get(_norm_pc(h.get("postcode", "")))
        h["miles"] = round(_haversine(origin, g), 1) if (origin and g) else None
        h["rank"] = rank_of(h)
        h["used_before"] = h["name"].lower() in known
    # Delali's order of approach: our own fleet, then tier 1, then tier 2.
    # Distance (and a haulier we've used on this lane) only breaks ties inside
    # a band - a closer tier-2 must never outrank a tier-1.
    ok.sort(key=lambda h: (h["rank"], h["miles"] is None,
                           h["miles"] or 9e9, not h["used_before"]))
    return (ok[:limit] if limit else ok), est


def main():
    a = sys.argv[1:]
    if a and a[0] == "import" and len(a) > 1:
        d = import_workbook(a[1])
        print(f"imported {len(d['hauliers'])} haulier(s), {len(d['couriers'])} courier(s), "
              f"{len(d['ctms'])} CTMS id(s) -> _hauliers.json")
    elif a and a[0] == "find":
        pc, needs, to = "", [], ""
        i = 1
        while i < len(a):
            if a[i] == "--pc": pc = a[i + 1]; i += 2
            elif a[i] == "--to": to = a[i + 1]; i += 2
            elif a[i] == "--need":
                i += 1
                while i < len(a) and not a[i].startswith("--"):
                    needs.append(a[i]); i += 1
            else: i += 1
        hits, est = recommend(pc, needs, to)
        if est:
            print(f"lane estimate: ~£{est['typical']:.0f} "
                  f"(£{est['price_low']:.0f}-£{est['price_high']:.0f}, {est['basis']})\n")
        for h in hits:
            m = f"{h['miles']}mi" if h["miles"] is not None else "  ? "
            tag = ("OUR FLEET", "TIER 1  ", "TIER 2  ")[h["rank"]]
            print(f"  {m:>7}  {tag}  {h['name'][:26]:26} {h['location'][:15]:15} "
                  f"{h['phone'][:22]:22} {'(used before)' if h['used_before'] else ''}")
            if h["emails"]: print(f"           {'; '.join(h['emails'][:2])}")
    elif a and a[0] == "show" and len(a) > 1:
        q = a[1].lower()
        d = load()
        for h in d.get("hauliers", []) + d.get("couriers", []) + d.get("services", []):
            if q in h["name"].lower():
                print(f"\n{h['name']}  [{h.get('ctms') or 'no CTMS id'}]")
                print(f"  {h['location']} {h['postcode']}  | tier alloc: {h['allocation']}")
                print(f"  day {h['phone']} | ooh {h['ooh']}")
                print(f"  {'; '.join(h['emails'])}")
                print(f"  can: {', '.join(h['caps'])}")
    else:
        print(__doc__)


if __name__ == "__main__":
    main()
